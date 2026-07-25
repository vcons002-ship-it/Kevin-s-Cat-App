"""Flask web GUI: the single page where Kevin configures and runs everything.

Serves the config page plus JSON endpoints:

  GET  /api/speakers   -> auto-detected Cast devices
  GET  /api/cameras    -> auto-detected ONVIF cameras
  GET  /api/sounds     -> sound files available to cast
  POST /api/sounds     -> upload a custom sound
  GET  /api/config     -> current saved settings
  POST /api/config     -> save settings
  POST /api/test       -> force a treat sound on the chosen speaker
  POST /api/start      -> start the detection loop
  POST /api/stop       -> stop the detection loop
  GET  /api/status     -> live loop status (running, last roll, counts)
  GET  /api/stream     -> live MJPEG feed (?camera=<name> selects which camera)
  GET  /api/diagnostics/detectors -> live detector settings vs the saved config
  POST /api/live/screenshot -> save the current picture of a camera, to keep
  POST /api/cameras/active -> set which saved cameras are watched at once
  GET  /api/cats       -> cat sightings (last seen, today's count, recent)
"""

from __future__ import annotations

import base64
import os
import threading
import time
import uuid
from collections import OrderedDict
from html import escape as esc

from flask import Flask, Response, jsonify, request, send_from_directory
from werkzeug.utils import secure_filename

from . import __version__
from . import config as config_mod
from . import discovery
from . import escalation
from . import moondream as vlm
from .cats import box_in_exit_zone, zone_for
from .detector import PersonDetector, grab_frame_jpeg, mask_credentials, sample_video_frames
from .loop import DetectionLoop, _camera_source

ALLOWED_SOUND_EXT = {".wav", ".mp3", ".ogg", ".m4a", ".aac"}

# Model provisioning (#86): one background run at a time; the GUI polls the log.
_PROVISION = {"running": False, "log": [], "error": None, "done": False}

# --- "Test detection" tool: upload a photo/video, run the net with adjustable
# settings, draw boxes. Frames are kept briefly in memory keyed by a session id so
# the GUI can re-run on each slider tweak without re-uploading. ---------------
_TEST_SESSIONS: "OrderedDict[str, list]" = OrderedDict()   # id -> [BGR frame, ...]
_TEST_SESSIONS_LOCK = threading.Lock()
# Must hold a whole batch's uploads at once — it's tied to the batch image ceiling
# (`_BENCH_MAX_IMAGES_HARD`) below so the two can't drift again: the old cap of 4 was
# sized for single-image testing and silently truncated big batches (#40), and the
# later 100 still forced full-set runs (199 cats + 43 nulls) to be split (#73).
# 1000 lets a whole benchmark set run in one pass; the models process images one at
# a time regardless, so this is upload-queue RAM only — NOT moondream's
# max_batch_size/kv_cache_pages VRAM params, which stay untouched (raising those
# OOMs the 8 GB card).
_TEST_MAX_SESSIONS = 1000
_TEST_VIDEO_FRAMES = 8           # fallback when a clip's duration can't be read
_TEST_MAX_FRAMES = 100           # cap video-frame extraction (#38)
_test_detectors: dict = {}              # (model, accelerator) -> reusable PersonDetector
_test_detect_lock = threading.Lock()    # serialise test detection + detector reuse
_TEST_MAX_UPLOAD = 256 * 1024 * 1024    # 256 MB cap on a test upload


def _thumb_data_url(frame, width: int = 200) -> str:
    import cv2

    h, w = frame.shape[:2]
    if w > width:
        frame = cv2.resize(frame, (width, max(1, int(h * width / w))))
    ok, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 70])
    if not ok:
        return ""
    return "data:image/jpeg;base64," + base64.b64encode(buf.tobytes()).decode("ascii")


def _video_default_frames(path: str) -> int:
    """~1 frame per second of clip (duration drives the count), clamped to
    ``[1, _TEST_MAX_FRAMES]``; falls back to ``_TEST_VIDEO_FRAMES`` when the
    container doesn't report fps/length (#38)."""
    import cv2

    cap = cv2.VideoCapture(path)
    try:
        fps = cap.get(cv2.CAP_PROP_FPS) or 0
        total = cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0
        if fps > 0 and total > 0:
            return max(1, min(_TEST_MAX_FRAMES, round(total / fps)))
    finally:
        cap.release()
    return _TEST_VIDEO_FRAMES


def _decode_test_upload(data: bytes, filename: str, n_frames=None) -> list:
    """Return BGR frames from uploaded image *or* video bytes ([] if unreadable).

    For a video, ``n_frames`` sets how many evenly-spaced frames to extract; when it
    is ``None`` the count defaults to ~1 fps of the clip's duration (#38). Always
    clamped to ``[1, _TEST_MAX_FRAMES]``.
    """
    import cv2
    import numpy as np

    img = cv2.imdecode(np.frombuffer(data, np.uint8), cv2.IMREAD_COLOR)
    if img is not None:
        return [img]
    # Not an image — try it as a video (cv2 needs a real path).
    import tempfile

    suffix = os.path.splitext(filename or "")[1] or ".mp4"
    tmp = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
    try:
        tmp.write(data)
        tmp.flush()
        tmp.close()
        n = n_frames if n_frames else _video_default_frames(tmp.name)
        n = max(1, min(_TEST_MAX_FRAMES, int(n)))
        return sample_video_frames(tmp.name, n)
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass


def _run_test_detection(frame, settings: dict, report: dict | None = None):
    """Run :meth:`PersonDetector.detect_image` with override settings.

    Reuses one detector per (model, accelerator) so the net isn't reloaded on every
    slider tweak; the cheap per-call knobs (confidence/size/floor/ROI/adjustments)
    are set per request under a lock.
    """
    model = settings.get("model") or "yolo11n"
    accel = settings.get("accelerator") or "cpu"

    def _f(key, default):
        try:
            return float(settings.get(key, default))
        except (TypeError, ValueError):
            return float(default)

    with _test_detect_lock:
        key = (model, accel)
        det = _test_detectors.get(key)
        if det is None:
            det = PersonDetector(source="__test__", model=model, accelerator=accel)
            _test_detectors[key] = det
        det.confidence = _f("person_confidence", 0.5)
        det.label_floor = _f("label_floor", 0.55)
        det.cat_confidence = _f("cat_confidence", 0.5)
        lc = settings.get("locator_classes")
        det.locator_classes = tuple(lc) if isinstance(lc, (list, tuple)) and lc else ("cat",)
        roi = settings.get("roi")
        det.roi = list(roi) if (isinstance(roi, (list, tuple)) and len(roi) == 4) else None
        det.gamma = _f("gamma", 1.0)
        det.brightness = int(_f("brightness", 0))
        det.contrast = _f("contrast", 1.0)
        det.saturation = _f("saturation", 1.0)
        det.cat_scan_tiling = settings.get("tiling") or "off"
        det.cat_scan_tile_overlap = _f("tile_overlap", 0.2)
        det.cat_scan_imgsz = int(_f("imgsz", 0))
        det._locator_tried = False        # re-resolve the larger-input net on setting change
        det._locator_runner = None
        if report is not None:
            # Read BACK off the detector, so this is what ran rather than what was
            # asked for. A setting the caller passed but nothing applied, or one a
            # cached detector kept from an earlier call, shows up here as a
            # difference — which is the only way to tell them apart from outside.
            eff, why = det.effective_accelerator()
            report.update({
                "model": det.model, "accelerator": accel,
                "ran_on": eff or accel, "fallback": why,
                "tiling": det.cat_scan_tiling, "tile_overlap": det.cat_scan_tile_overlap,
                "grid": det._tiling_grid(), "imgsz": det.cat_scan_imgsz,
                "scan_model": det.cat_scan_model,
                "cat_confidence": det.cat_confidence, "label_floor": det.label_floor,
                "person_confidence": det.confidence,
                "locator_classes": list(det.locator_classes),
                "roi": det.roi, "frame_size": [frame.shape[1], frame.shape[0]],
                "gamma": det.gamma, "brightness": det.brightness,
                "contrast": det.contrast, "saturation": det.saturation,
            })
        t0 = time.perf_counter()
        annotated, dets = det.detect_image(frame)
        return annotated, dets, round((time.perf_counter() - t0) * 1000.0, 1)


def _effective_accel(model: str, accel: str):
    """What the cached test detector for (model, accel) actually ran on (#90):
    ``(effective, reason)`` — or ``(None, "")`` before its first run."""
    with _test_detect_lock:
        det = _test_detectors.get((model, accel))
    return det.effective_accelerator() if det is not None else (None, "")


def _yolo_boxes_fn(settings: dict, floor: float):
    """A ``run_yolo(img) -> [(label, score, box)]`` callable for the escalation
    ladder (#66), reusing the Test tool's per-(model, accelerator) detector cache
    under the same lock (cv2.dnn nets aren't thread-safe). Defaults to **CPU** —
    the ladder is on-demand and the 8 GB card may already hold moondream (~4.6 GB);
    pass ``settings["accelerator"]`` to override."""
    model = settings.get("model") or "yolo11n"
    accel = settings.get("accelerator") or "cpu"

    def run_yolo(img):
        with _test_detect_lock:
            key = (model, accel)
            det = _test_detectors.get(key)
            if det is None:
                det = PersonDetector(source="__test__", model=model, accelerator=accel)
                _test_detectors[key] = det
            return det._run_net(img, floor)

    return run_yolo


# --- "Benchmark this image": sweep models × tiling on one frame, emit a shareable
# self-contained HTML report + an optional XLSX (needs openpyxl). ---------------
_BENCHMARKS: "OrderedDict[str, dict]" = OrderedDict()     # id -> {html, xlsx, xlsx_error}
_BENCH_LOCK = threading.Lock()
_BENCH_MAX_REPORTS = 40         # hold a full batch (per-image reports + summary) at once
_BENCH_TILINGS = ["off", "2x2", "3x3", "4x4"]
_BENCH_MAX_RUNS = 24            # cap the matrix so one request can't run forever
# /guide: a deliberately tiny markdown-to-HTML converter — headings, bold/italic,
# inline + fenced code, lists, links, hr. Enough for docs/USER_GUIDE.md; NOT a
# general renderer (tables/images/nesting unsupported — keep the guide simple).
_GUIDE_STYLE = (
    "<!DOCTYPE html><html><head><meta charset='utf-8'>"
    "<meta name='viewport' content='width=device-width, initial-scale=1'>"
    "<title>Kevin's Cat App — User Guide</title><style>"
    "body{font-family:system-ui,sans-serif;max-width:860px;margin:2rem auto;"
    "padding:0 1rem;line-height:1.55;color:#222;background:#fafafa}"
    "h1,h2{border-bottom:1px solid #ddd;padding-bottom:.3rem}"
    "code{background:#eee;padding:1px 5px;border-radius:4px;font-size:.92em}"
    "pre{background:#222;color:#eee;padding:.8rem;border-radius:8px;overflow-x:auto}"
    "pre code{background:none;color:inherit;padding:0}"
    "a{color:#1565c0}hr{border:none;border-top:1px solid #ddd;margin:1.5rem 0}"
    "li{margin:.25rem 0}"
    "@media(prefers-color-scheme:dark){body{background:#16181c;color:#ddd}"
    "h1,h2{border-color:#333}code{background:#2a2d33}a{color:#7fb2ff}"
    "hr{border-color:#333}}"
    "</style></head><body>")


def _md_to_html(md: str) -> str:
    import html as html_mod
    import re as re_mod

    def inline(s):
        s = html_mod.escape(s, quote=False)
        s = re_mod.sub(r"`([^`]+)`", r"<code>\1</code>", s)
        s = re_mod.sub(r"\*\*([^*]+)\*\*", r"<strong>\1</strong>", s)
        s = re_mod.sub(r"(?<!\*)\*([^*]+)\*(?!\*)", r"<em>\1</em>", s)
        s = re_mod.sub(r"\[([^\]]+)\]\(([^)\s]+)\)", r"<a href='\2'>\1</a>", s)
        return s

    out, para, in_list, in_code = [], [], False, False

    def flush_para():
        nonlocal para
        if para:
            out.append("<p>" + " ".join(inline(x) for x in para) + "</p>")
            para = []

    def close_list():
        nonlocal in_list
        if in_list:
            out.append("</ul>")
            in_list = False

    for line in md.splitlines():
        if line.strip().startswith("```"):
            flush_para(); close_list()
            out.append("<pre><code>" if not in_code else "</code></pre>")
            in_code = not in_code
            continue
        if in_code:
            out.append(html_mod.escape(line))
            continue
        m = re_mod.match(r"^(#{1,4})\s+(.*)$", line)
        if m:
            flush_para(); close_list()
            n = len(m.group(1))
            out.append(f"<h{n}>{inline(m.group(2))}</h{n}>")
            continue
        if re_mod.match(r"^---+\s*$", line):
            flush_para(); close_list()
            out.append("<hr>")
            continue
        m = re_mod.match(r"^\s*[-*]\s+(.*)$", line)
        if m:
            flush_para()
            if not in_list:
                out.append("<ul>")
                in_list = True
            out.append(f"<li>{inline(m.group(1))}</li>")
            continue
        if not line.strip():
            flush_para(); close_list()
            continue
        # a continuation line inside a list item joins that item
        if in_list and out and out[-1].endswith("</li>"):
            out[-1] = out[-1][:-5] + " " + inline(line.strip()) + "</li>"
            continue
        para.append(line.strip())
    flush_para(); close_list()
    if in_code:
        out.append("</code></pre>")
    return "\n".join(out) + "</body></html>"


def _model_options() -> list:
    """``[{value, label}]`` for every selectable model, straight from the registry
    (labels + the ``selectable`` flag live in ``yolo.MODELS`` — one source, no drift;
    #50, #71). Only present ONNX files are offered (export-only ones like yolo26x /
    the fp16 variants appear once their file exists), and #71's dropped models
    (11m and friends) are registered-but-not-selectable: old configs load, new
    configs can't pick them. Order = the registry's lightest-to-heaviest."""
    from . import provision, yolo

    # #90: precision is not a user choice — the picker shows the 3 LOGICAL
    # models; resolve_variant() maps (model, accelerator) to the FP32/FP16 file.
    # A model is offered when either precision's file exists (the accelerator
    # in use decides which one actually loads).
    # #86: a present file the manifest can't vouch for (unknown provenance or
    # changed since vetting) is offered but SAYS SO — never silently run a
    # model that may not be the settled precision/head.
    flags = {r["file"]: r["status"] for r in provision.audit()
             if r["kind"] == "onnx" and r["status"] in ("unverified", "stale")}
    out = []
    for v, m in yolo.MODELS.items():
        if not m.get("selectable", True):
            continue
        files = [m["file"]]
        fp16 = yolo.MODELS.get(f"{v}_fp16")
        if fp16:
            files.append(fp16["file"])
        if not any(os.path.exists(os.path.join(
                os.path.dirname(yolo.model_path(v)), f)) for f in files):
            continue
        label = m.get("label", v)
        # #106: keep the flag SHORT so the dropdown doesn't overflow the app
        # width — the detail lives in the Model files card.
        statuses = {flags[f] for f in files if f in flags}
        if statuses:
            label += f" ⚠ {'/'.join(sorted(statuses))}"
        out.append({"value": v, "label": label})
    return out


def _benchmark_models():
    """The model values a sweep offers — derived from :func:`_model_options` so the
    sweep and the GUI dropdowns share one registry-backed source and can't drift."""
    return [m["value"] for m in _model_options()]


def _model_native_size(model: str) -> int:
    from . import yolo

    if model in yolo.MODELS:
        return yolo.input_size(model)
    return 0


def _display_model(model: str) -> str:
    """Visible model name for reports. All models are now YOLO variants whose size is
    their export, so the name stands on its own."""
    return model


def _slugify(text: str) -> str:
    """Lowercase, hyphenated slug for a human-readable report filename."""
    import re

    s = re.sub(r"[^a-z0-9]+", "-", (text or "").lower()).strip("-")
    return s[:60] or "report"


def _full_jpeg_data_url(frame, quality: int = 95) -> str:
    """The unannotated frame at full resolution, as a JPEG data URL — embedded once
    in the report so the exact input can be re-run (reproducibility)."""
    import cv2

    ok, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, quality])
    return ("data:image/jpeg;base64," + base64.b64encode(buf.tobytes()).decode("ascii")) if ok else ""


def _bench_thumb(jpeg: bytes, width: int = 680):
    """(data-url, raw-jpeg-bytes) thumbnail of an annotated frame for the reports.

    Encoded at ~680px (not the old 240) so the drawn boxes/labels are actually
    legible when enlarged; the report still displays them small but click-to-zoom.
    """
    import cv2
    import numpy as np

    img = cv2.imdecode(np.frombuffer(jpeg, np.uint8), cv2.IMREAD_COLOR)
    if img is None:
        return "", b""
    h, w = img.shape[:2]
    if w > width:
        img = cv2.resize(img, (width, max(1, int(h * width / w))))
    ok, buf = cv2.imencode(".jpg", img, [cv2.IMWRITE_JPEG_QUALITY, 72])
    raw = buf.tobytes() if ok else b""
    return ("data:image/jpeg;base64," + base64.b64encode(raw).decode("ascii")), raw


# The tester controls the benchmark now applies uniformly to every run (#44), with
# the Test-detection defaults. The cat/locator score is still captured at a low floor
# (so the score column stays meaningful); the user's cat threshold marks "found?".
_BENCH_OPT_DEFAULTS = {"tile_overlap": 0.2, "person_confidence": 0.5,
                       "gamma": 1.0, "brightness": 0, "contrast": 1.0, "saturation": 1.0}


def _bench_opts(body: dict) -> dict:
    """Pull the tester controls from a request, falling back to the tester defaults so
    existing behaviour is unchanged when they're omitted (#44)."""
    out = {}
    for k, d in _BENCH_OPT_DEFAULTS.items():
        try:
            out[k] = (int(body.get(k, d)) if k == "brightness" else float(body.get(k, d)))
        except (TypeError, ValueError):
            out[k] = d
    return out


def _run_benchmark(frame, models: list, tilings: list, cat_threshold: float,
                   accelerator: str, opts: dict = None) -> list:
    """Sweep ``models × tilings`` on one frame. Each run reuses ``_run_test_detection``
    (no parallel detection path) with a very low floor + cat/dog locator so the
    **raw** best cat/dog score is captured; ``detected`` marks it against
    ``cat_threshold``. The tester controls in ``opts`` (tile overlap, person
    confidence, gamma/brightness/contrast/saturation) are applied uniformly to every
    run (#44). Sorted by combined (cat-or-dog) score, best first."""
    opts = {**_BENCH_OPT_DEFAULTS, **(opts or {})}
    runs = []
    for model in models:
        for tiling in tilings:
            settings = {
                "model": model, "tiling": tiling,
                "tile_overlap": opts["tile_overlap"],
                "accelerator": accelerator,
                "person_confidence": opts["person_confidence"],
                "gamma": opts["gamma"], "brightness": opts["brightness"],
                "contrast": opts["contrast"], "saturation": opts["saturation"],
                "cat_confidence": 0.01, "label_floor": 0.01,
                "locator_classes": ["cat", "dog"],
            }
            annotated, dets, ms = _run_test_detection(frame, settings)
            best_cat = max((d["score"] for d in dets if d["label"] == "cat"), default=0.0)
            best_dog = max((d["score"] for d in dets if d["label"] == "dog"), default=0.0)
            combined = max(best_cat, best_dog)
            thumb, raw = _bench_thumb(annotated)
            eff, why = _effective_accel(model, accelerator)
            runs.append({
                "model": model, "size": _model_native_size(model), "tiling": tiling,
                "tile_overlap": opts["tile_overlap"], "accelerator": accelerator,
                # #90: what ACTUALLY ran — a fallback must be visible in the
                # report, never its numbers silently labelled as the request.
                "ran_on": eff or accelerator,
                "fallback": why,
                "cat_score": round(best_cat, 3), "dog_score": round(best_dog, 3),
                "combined_score": round(combined, 3),
                "detected": bool(combined >= cat_threshold),
                "boxes": sum(1 for d in dets if d["label"] in ("cat", "dog")),
                "inference_ms": ms, "thumb": thumb, "_raw": raw,
            })
    runs.sort(key=lambda r: -r["combined_score"])
    return runs


def _fixed_settings_text(meta: dict) -> str:
    """The 'held fixed for every run' line for a report — the **actual** values used,
    not a hardcoded string, so two runs under different settings are distinguishable
    on paper (#44). Non-default image adjustments are appended only when set."""
    o = {**_BENCH_OPT_DEFAULTS,
         **{k: meta[k] for k in _BENCH_OPT_DEFAULTS if meta.get(k) is not None}}
    parts = [f"cat threshold {meta['cat_threshold']:.2f}",
             f"accelerator {meta['accelerator']}",
             f"tile overlap {float(o['tile_overlap']):.2f}",
             f"person {float(o['person_confidence']):.2f}",
             "locator [cat, dog]"]
    adj = []
    if abs(float(o["gamma"]) - 1.0) > 1e-9:
        adj.append(f"gamma {float(o['gamma']):.2f}")
    if int(o["brightness"]) != 0:
        adj.append(f"brightness {int(o['brightness']):+d}")
    if abs(float(o["contrast"]) - 1.0) > 1e-9:
        adj.append(f"contrast {float(o['contrast']):.2f}")
    if abs(float(o["saturation"]) - 1.0) > 1e-9:
        adj.append(f"saturation {float(o['saturation']):.2f}")
    if adj:
        parts.append("adjust: " + ", ".join(adj))
    return " · ".join(parts)


def _score_color(score: float) -> str:
    """Red→amber→green for a 0..1 score (report cell background)."""
    s = max(0.0, min(1.0, float(score)))
    r, g = (220, int(60 + 160 * s)) if s < 0.5 else (int(220 - 320 * (s - 0.5)), 200)
    return f"rgb({r},{g},70)"


def _benchmark_html(runs: list, meta: dict, original_url: str = "",
                    original_name: str = "original.jpg") -> str:
    """A self-contained HTML report — every thumbnail inlined, no external assets,
    so it renders for a remote viewer and emails cleanly. Each run thumbnail is
    click-to-enlarge, and the full-resolution **original** frame is embedded once at
    the top with a download link, so the exact test can be re-run later."""
    rows = []
    for r in runs:
        hit = "✓" if r["detected"] else "·"
        rows.append(
            f"<tr><td><img src='{r['thumb']}' title='click to enlarge' "
            f"onclick='zoom(this.src)'></td>"
            f"<td>{esc(_display_model(r['model']))}</td><td>{r['size']}</td>"
            f"<td>{esc(r['tiling'])}</td>"
            f"<td style='background:{_score_color(r['combined_score'])}'>"
            f"<b>{r['combined_score']:.2f}</b></td>"
            f"<td>{r['cat_score']:.2f}</td><td>{r['dog_score']:.2f}</td>"
            f"<td>{hit}</td><td>{r['inference_ms']:.0f} ms</td></tr>")
    fixed = esc(_fixed_settings_text(meta))
    original = ""
    if original_url:
        # Enlarge in-page via the lightbox (browsers block top-level navigation to
        # a data: URL — #30); the download link keeps the `download` attribute,
        # which still works for data: URLs.
        original = (
            "<details open><summary><b>Original frame</b> "
            f"({meta['width']}×{meta['height']}) — "
            f"<a href='{original_url}' download='{esc(original_name)}'>download to re-run</a>"
            "</summary>"
            f"<img class='orig' src='{original_url}' title='click to enlarge' "
            f"onclick='zoom(this.src)'></details>")
    return (
        "<!doctype html><html><head><meta charset='utf-8'>"
        "<title>Cat detection benchmark</title><style>"
        "body{font-family:system-ui,Arial,sans-serif;margin:24px;color:#111}"
        "table{border-collapse:collapse;width:100%}th,td{border:1px solid #ddd;"
        "padding:6px 8px;text-align:center;font-size:14px}th{background:#f3f3f3}"
        "td img{max-width:240px;height:auto;border-radius:4px;display:block;cursor:zoom-in}"
        "img.orig{max-width:100%;border-radius:6px;margin:8px 0;cursor:zoom-in}"
        ".fixed{background:#f7f7f9;border:1px solid #e3e3e8;border-radius:8px;"
        "padding:10px 14px;margin:10px 0;font-size:14px}"
        "#lb{position:fixed;inset:0;background:rgba(0,0,0,.85);display:none;"
        "align-items:center;justify-content:center;cursor:zoom-out;z-index:9}"
        "#lb img{max-width:96vw;max-height:96vh;border-radius:6px}"
        "</style></head><body>"
        # A tiny self-contained lightbox: set its <img> src on click, hide on click.
        "<div id='lb' onclick='this.style.display=\"none\"'><img id='lbimg'></div>"
        "<script>function zoom(s){var b=document.getElementById('lb');"
        "document.getElementById('lbimg').src=s;b.style.display='flex';}</script>"
        f"<h2>🐱 Cat detection benchmark — {esc(meta['image'])}</h2>"
        f"<p class='muted'>{esc(meta['ts'])} · {len(runs)} runs</p>"
        f"<div class='fixed'><b>Settings held fixed for every run:</b> {fixed}</div>"
        + original +
        "<table><tr><th>frame</th><th>model</th><th>size</th><th>tiling</th>"
        "<th>cat-or-dog</th><th>cat</th><th>dog</th><th>found?</th><th>time</th></tr>"
        + "".join(rows) + "</table>"
        "<p style='color:#777;font-size:13px'>“cat-or-dog” is the combined locator "
        "score (the number that predicts locator reliability with cat+dog); “found?” "
        "marks it against the cat threshold above. Time is the total for the run "
        "(summed across tiles).</p></body></html>")


def _benchmark_xlsx(runs: list, meta: dict) -> bytes:
    """XLSX with embedded thumbnails. Raises if openpyxl isn't installed."""
    try:
        import openpyxl
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font
    except Exception as exc:        # noqa: BLE001 — optional dependency
        raise RuntimeError(
            "XLSX export needs the 'openpyxl' package — re-run setup (say yes to "
            "spreadsheet export) or: pip install openpyxl") from exc
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "benchmark"
    ws["A1"] = f"Cat detection benchmark — {meta['image']} ({meta['ts']})"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "Fixed: " + _fixed_settings_text(meta)
    headers = ["frame", "model", "size", "tiling", "cat-or-dog", "cat", "dog",
               "found?", "time (ms)"]
    ws.append([])
    ws.append(headers)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    head_row = ws.max_row
    for i, r in enumerate(runs):
        row = head_row + 1 + i
        ws.cell(row, 2, _display_model(r["model"])); ws.cell(row, 3, r["size"])
        ws.cell(row, 4, r["tiling"]); ws.cell(row, 5, r["combined_score"])
        ws.cell(row, 6, r["cat_score"]); ws.cell(row, 7, r["dog_score"])
        ws.cell(row, 8, "yes" if r["detected"] else "no")
        ws.cell(row, 9, r["inference_ms"])
        ws.row_dimensions[row].height = 70
        if r.get("_raw"):
            img = XLImage(io.BytesIO(r["_raw"]))
            img.width, img.height = 120, 90
            ws.add_image(img, f"A{row}")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Batch benchmark: run the sweep across many images and aggregate which config
# is most reliable across all of them (the cross-image question single reports can't
# answer). Per-image reports are still produced; the summary links out to them. ----
_BENCH_MAX_IMAGES = 12          # *recommended* batch size — the GUI warns past this
# Absolute ceiling (soft cap above; #37). Tied to the upload-session cap so a batch can
# never accept more images than the session store can hold at once (#40).
_BENCH_MAX_IMAGES_HARD = _TEST_MAX_SESSIONS
_BENCH_CANCEL: set = set()      # batch ids the client asked to abort mid-run (#37)
_BENCH_CANCEL_LOCK = threading.Lock()


def _orig_thumb_data_url(frame, width: int = 200) -> str:
    """A small **unannotated** original thumbnail (data URL) for the summary's image
    catalog — embedded once per image (not once per miss), so summary size scales
    with image count, not miss count (#32 3c)."""
    import cv2

    h, w = frame.shape[:2]
    img = cv2.resize(frame, (width, max(1, int(h * width / w)))) if w > width else frame
    ok, buf = cv2.imencode(".jpg", img, [cv2.IMWRITE_JPEG_QUALITY, 70])
    return ("data:image/jpeg;base64," + base64.b64encode(buf.tobytes()).decode("ascii")) if ok else ""


def _aggregate_configs(per_image: list) -> list:
    """Aggregate every model×tiling config across all images. For cat-present images:
    detection rate, average combined score, and which images it missed. For no-cat
    images: how often it false-fired. Average inference time over all images. Sorted
    by detection rate, then average score (the deploy-config shortlist)."""
    cat_imgs = [im for im in per_image if im["has_cat"]]
    nocat_imgs = [im for im in per_image if not im["has_cat"]]
    keys = OrderedDict()        # preserve first-seen order
    for im in per_image:
        for r in im["runs"]:
            keys.setdefault((r["model"], r["tiling"]), r["size"])

    def run_for(im, model, tiling):
        return next((r for r in im["runs"]
                     if r["model"] == model and r["tiling"] == tiling), None)

    out = []
    for (model, tiling), size in keys.items():
        cat_runs = [(im, run_for(im, model, tiling)) for im in cat_imgs]
        found = [im["idx"] for im, r in cat_runs if r and r["detected"]]
        misses = [im["idx"] for im, r in cat_runs if r and not r["detected"]]
        scores = [r["combined_score"] for _, r in cat_runs if r]
        all_ms = [r["inference_ms"] for r in
                  (run_for(im, model, tiling) for im in per_image) if r]
        fp = [im["idx"] for im in nocat_imgs
              if (run_for(im, model, tiling) or {}).get("detected")]
        out.append({
            "model": model, "tiling": tiling, "size": size,
            "found": len(found), "total": len(cat_imgs),
            "rate": (len(found) / len(cat_imgs)) if cat_imgs else 0.0,
            "avg_score": round(sum(scores) / len(scores), 3) if scores else 0.0,
            "avg_ms": round(sum(all_ms) / len(all_ms), 1) if all_ms else 0.0,
            "misses": misses, "fp": len(fp),
            "fp_total": len(nocat_imgs), "fp_imgs": fp,
        })
    out.sort(key=lambda c: (-c["rate"], -c["avg_score"]))
    return out


def _vlm_block_html(vlm_info: dict, per_image: list) -> str:
    """The VLM section of a cross-image summary (#54): one accuracy line (recall + FP,
    separate) and the frames where the VLM and the best YOLO config disagree — the
    interesting rows. Empty string when the toggle didn't run."""
    if not vlm_info:
        return ""
    if vlm_info.get("error"):
        return ("<div class='fixed'><b>VLM:</b> skipped — "
                f"{esc(vlm_info['error'])}</div>")
    s = vlm_info["summary"]
    by_slug = {im["slug"]: im for im in per_image}
    rec = "—" if s["recall"] is None else f"{s['recall'] * 100:.0f}% ({s['found']}/{s['n_cat']})"
    fpr = "—" if s["fp_rate"] is None else f"{s['fp_rate'] * 100:.0f}% ({s['fp']}/{s['n_nocat']})"
    dis = vlm_info.get("disagreements") or []
    passes = vlm_info.get("passes") or 1
    rows = ""
    for d in dis:
        im = by_slug.get(d["slug"], {})
        thumb = (f"<img src='{im.get('orig_thumb', '')}' onclick='zoom(this.src)' "
                 "title='click to enlarge'>") if im.get("orig_thumb") else ""
        gt = "cat" if d["has_cat"] else "no-cat"
        ratio = f" ({esc(d['ratio'])})" if d.get("ratio") else ""
        rows += (f"<tr><td>{thumb}</td><td style='text-align:left'>"
                 f"<a href='{d['slug']}.html' target='_blank'>{esc(d['name'])}</a></td>"
                 f"<td>{esc(gt)}</td><td>VLM {esc(d['vlm'])}{ratio}</td>"
                 f"<td>YOLO {esc(d['yolo'])}</td></tr>")
    dis_html = ("<p class='muted' style='font-size:13px'>VLM and the best YOLO config "
                "agree on every frame.</p>" if not dis else
                "<table><tr><th>frame</th><th>image</th><th>truth</th><th></th><th></th>"
                f"</tr>{rows}</table>")
    # Borderline (non-unanimous across the N passes) — the genuinely hard frames (#60).
    bl = vlm_info.get("borderline") or []
    bl_html = ("" if passes <= 1 else
               ("<p class='muted' style='font-size:13px'>Every frame's vote was "
                "unanimous across the passes.</p>" if not bl else
                "<table><tr><th>image</th><th>verdict</th><th>vote</th></tr>" +
                "".join(f"<tr><td style='text-align:left'>"
                        f"<a href='{b['slug']}.html' target='_blank'>{esc(b['name'])}</a></td>"
                        f"<td>{esc(b['answer'] or 'tie')}</td><td>{esc(b['ratio'])}</td></tr>"
                        for b in bl) + "</table>"))
    passes_note = "" if passes <= 1 else (
        f" <span class='muted'>· {passes} passes/image, majority vote (the vote ratio is "
        "the confidence)</span>")
    bl_section = "" if passes <= 1 else (
        "<h3>Borderline frames (split vote — review)</h3>" + bl_html)
    return (
        f"<div class='fixed'><b>VLM ({esc(vlm_info['model'])}) accuracy for batch:</b> "
        f"recall <b>{rec}</b> · false-positive rate <b>{fpr}</b> "
        "<span class='muted'>(recall on cat frames, FP on no-cat frames — reported "
        f"separately)</span>{passes_note}</div>"
        "<h3>VLM ⇄ YOLO disagreements</h3>" + dis_html + bl_section)


def _benchmark_summary_html(per_image: list, configs: list, meta: dict, vlm=None) -> str:
    """The cross-image summary: a clean config table (sorted by detection rate) whose
    'found' cell expands to the missed frames, a config×image heatmap, and an image
    catalog embedded once each. Links out to the per-image reports (kept small)."""
    n_cat = sum(1 for im in per_image if im["has_cat"])
    n_nocat = len(per_image) - n_cat
    # Image catalog: one embedded original thumbnail per image (by idx), for the
    # miss lists and the lightbox — referenced, never duplicated per miss.
    cat_js = ",".join(f"'{im['orig_thumb']}'" for im in per_image)
    by_idx = {im["idx"]: im for im in per_image}

    def miss_detail(c):
        if not c["misses"]:
            return ""
        cells = []
        for idx in c["misses"]:
            im = by_idx[idx]
            r = next((x for x in im["runs"]
                      if x["model"] == c["model"] and x["tiling"] == c["tiling"]), None)
            sc = f"{r['combined_score']:.2f}" if r else "—"
            # Link by the report's SLUG (not its hash id): that's the filename each
            # per-image report downloads under, so a downloaded/hosted bundle stays
            # self-consistent. The serving route resolves a slug too (#35).
            # Lead with the image number so the grid column (#N) and the miss list use
            # the same identifier (#42); link by slug (matches the download filename).
            cells.append(
                f"<div class='miss'><img src='{im['orig_thumb']}' onclick='zoom(this.src)' "
                f"title='click to enlarge'><div><a href='{im['slug']}.html' "
                f"target='_blank'>#{idx + 1} {esc(im['name'])}</a><br>"
                f"<span class='sub'>scored {sc}, below {meta['cat_threshold']:.2f}</span>"
                "</div></div>")
        return "<div class='misses'>" + "".join(cells) + "</div>"

    rows, details = [], []
    for i, c in enumerate(configs):
        label = f"{esc(_display_model(c['model']))} {c['size']} {esc(c['tiling'])}"
        rate_txt = f"{c['found']}/{c['total']}"
        perfect = c["found"] == c["total"]
        # Only an imperfect rate with traceable misses invites a click.
        if c["misses"]:
            found_cell = (f"<td class='exp' onclick='tog({i})'>{rate_txt} "
                          f"<span class='car'>▸</span></td>")
        else:
            mark = "✓" if perfect and c["total"] else ""
            found_cell = f"<td>{rate_txt} {mark}</td>"
        fp_cell = (f"{c['fp']}/{c['fp_total']}" if n_nocat else "—")
        rows.append(
            f"<tr><td style='text-align:left'>{label}</td>{found_cell}"
            f"<td style='background:{_score_color(c['avg_score'])}'>"
            f"{c['avg_score']:.2f}</td><td>{c['avg_ms']:.0f} ms</td>"
            f"<td>{fp_cell}</td></tr>")
        if c["misses"]:
            details.append(
                f"<tr id='d{i}' class='detail' style='display:none'><td colspan='5'>"
                f"<b>Missed:</b> {miss_detail(c)}</td></tr>")
        else:
            details.append("")
    body_rows = "".join(r + d for r, d in zip(rows, details))

    # config × image heatmap: rows = configs (best first), cols = images. Each column
    # header links to that image's per-image report (by slug) and reveals its name +
    # thumbnail on hover, so "column #5 is hard" is one click from frame #5's report (#42).
    head_cells = "".join(
        f"<th class='imgcol'><a href='{im['slug']}.html' target='_blank' "
        f"title='{esc(im['name'])}'>#{im['idx'] + 1}"
        + ("" if im["has_cat"] else " ∅") + "</a>"
        f"<span class='tip'><img src='{im['orig_thumb']}'><br>{esc(im['name'])}</span>"
        "</th>" for im in per_image)
    grid_rows = []
    for c in configs:
        cells = []
        for im in per_image:
            r = next((x for x in im["runs"] if x["model"] == c["model"]
                      and x["tiling"] == c["tiling"]), None)
            sc = r["combined_score"] if r else 0.0
            if im["has_cat"]:
                txt = f"{sc:.2f}" if (r and r["detected"]) else "✗"
                bg = _score_color(sc) if (r and r["detected"]) else "#e66"
            else:                       # no-cat control: a detection here is a false +
                txt = "FP" if (r and r["detected"]) else "·"
                bg = "#e66" if (r and r["detected"]) else "#eee"
            cells.append(f"<td title='{esc(im['name'])}: {sc:.2f}' "
                         f"style='background:{bg}'>{txt}</td>")
        grid_rows.append(
            f"<tr><td style='text-align:left'>{esc(_display_model(c['model']))} "
            f"{c['size']} {esc(c['tiling'])}</td>" + "".join(cells) + "</tr>")

    fp_note = (f" · {n_nocat} no-cat control"
               f"{'s' if n_nocat != 1 else ''} (false-positive check)"
               if n_nocat else "")
    return (
        "<!doctype html><html><head><meta charset='utf-8'>"
        "<title>Cat detection benchmark — cross-image summary</title><style>"
        "body{font-family:system-ui,Arial,sans-serif;margin:24px;color:#111}"
        "h2{margin-bottom:2px}table{border-collapse:collapse;margin:12px 0}"
        "th,td{border:1px solid #ddd;padding:6px 9px;text-align:center;font-size:14px}"
        "th{background:#f3f3f3}td.exp{cursor:pointer;user-select:none}"
        ".car{color:#888;font-size:12px}.detail td{text-align:left;background:#fafafa}"
        ".misses{display:flex;flex-wrap:wrap;gap:12px;margin:6px 0}"
        ".miss{display:flex;gap:8px;align-items:center;font-size:13px}"
        ".miss img{width:90px;border-radius:4px;cursor:zoom-in}"
        ".miss .sub{color:#a33}.grid td{font-size:12px;padding:4px 6px}"
        "th.imgcol{position:relative}th.imgcol a{color:#06c;text-decoration:none}"
        "th.imgcol .tip{display:none;position:absolute;z-index:6;top:100%;left:50%;"
        "transform:translateX(-50%);background:#fff;border:1px solid #bbb;border-radius:4px;"
        "padding:4px;font-size:11px;white-space:nowrap;box-shadow:0 2px 6px rgba(0,0,0,.2)}"
        "th.imgcol:hover .tip{display:block}th.imgcol .tip img{width:150px;display:block}"
        ".fixed{background:#f7f7f9;border:1px solid #e3e3e8;border-radius:8px;"
        "padding:10px 14px;margin:10px 0;font-size:14px}"
        "#lb{position:fixed;inset:0;background:rgba(0,0,0,.85);display:none;"
        "align-items:center;justify-content:center;cursor:zoom-out;z-index:9}"
        "#lb img{max-width:96vw;max-height:96vh;border-radius:6px}"
        "</style></head><body>"
        "<div id='lb' onclick='this.style.display=\"none\"'><img id='lbimg'></div>"
        f"<script>var CAT=[{cat_js}];"
        "function zoom(s){var b=document.getElementById('lb');"
        "document.getElementById('lbimg').src=s;b.style.display='flex';}"
        "function tog(i){var r=document.getElementById('d'+i);"
        "if(!r)return;r.style.display=r.style.display==='none'?'table-row':'none';}"
        "</script>"
        "<h2>🐱 Cat detection benchmark — cross-image summary</h2>"
        f"<p class='muted'>{esc(meta['ts'])} · {len(per_image)} images "
        f"({n_cat} with a cat{fp_note}) · {len(configs)} configs</p>"
        f"<div class='fixed'><b>Held fixed for every run:</b> "
        f"{esc(_fixed_settings_text(meta))}. "
        "<b>found</b> = images where the cat cleared the threshold; click an imperfect "
        "rate to see which frames it missed.</div>"
        + _vlm_block_html(vlm, per_image) +
        "<table><tr><th style='text-align:left'>config</th><th>found</th>"
        "<th>avg conf</th><th>avg ms</th><th>false +</th></tr>"
        + body_rows + "</table>"
        "<h3>config × image</h3>"
        "<p class='muted' style='font-size:13px'>✗ = cat missed · FP = fired on a "
        "no-cat frame (∅) · number = combined score. Hover a cell for the image.</p>"
        "<table class='grid'><tr><th style='text-align:left'>config</th>"
        + head_cells + "</tr>" + "".join(grid_rows) + "</table>"
        "</body></html>")


def _find_report(key: str):
    """Resolve a stored report by its hash id **or** its slug, so the slug-based links
    a downloaded/hosted bundle uses also resolve on-server (#35)."""
    rep = _BENCHMARKS.get(key)
    if rep:
        return rep
    return next((r for r in _BENCHMARKS.values() if r.get("slug") == key), None)


# --- VLM (moondream) batch: one query per image, scored against the SAME "cat present"
# ground-truth labels as the detection benchmark, with recall and false-positive rate
# reported separately (#54). The VLM runs once per image — no tiling, no per-config sweep.
def _vlm_summary(verdicts: list) -> dict:
    """Recall on cat-present frames and false-positive rate on no-cat frames, kept
    separate (same methodology as the detection benchmark). ``None`` rates when there
    are no frames of that kind; errored frames are excluded from both."""
    cat = [v for v in verdicts if v.get("has_cat") and not v.get("error")]
    nocat = [v for v in verdicts if not v.get("has_cat") and not v.get("error")]
    found = sum(1 for v in cat if v.get("answer") == "yes")
    fp = sum(1 for v in nocat if v.get("answer") == "yes")
    return {
        "n_cat": len(cat), "found": found,
        "recall": round(found / len(cat), 3) if cat else None,
        "n_nocat": len(nocat), "fp": fp,
        "fp_rate": round(fp / len(nocat), 3) if nocat else None,
        "errors": sum(1 for v in verdicts if v.get("error")),
    }


def _run_vlm_items(items: list, model: str, batch_id=None, mode="local", api_key=None,
                   passes=1, prompt=None, reasoning=False):
    """Run a majority-voted VLM query per uploaded item; return ``(verdicts, skipped,
    cancelled)``. Each verdict carries ``name``/``has_cat``/``answer`` (the majority
    verdict)/``reason``/``ratio``/``unanimous``/``borderline``/``query_ms``/``error``.
    ``prompt`` is the USER'S prompt, threaded through exactly like the single-image
    path (#72 — the batch used to silently run the baked-in default, making batch
    prompt-testing impossible); None falls back to the model's validated default.
    Cancel- and skip-aware, mirroring the detection batch. Caller should ``preflight``
    first so a setup error (no package/key/GPU) fails fast rather than once per image."""
    prompt = (prompt or "").strip() or vlm.default_prompt(model)
    verdicts, skipped = [], []
    cancelled = False
    for i, it in enumerate(items):
        if batch_id:
            with _BENCH_CANCEL_LOCK:
                if batch_id in _BENCH_CANCEL:
                    cancelled = True
        if cancelled:
            break
        it = it or {}
        name = it.get("name") or f"image {i + 1}"
        has_cat = it.get("has_cat", True) is not False
        with _TEST_SESSIONS_LOCK:
            frames = _TEST_SESSIONS.get(it.get("id"))
        if not frames:
            skipped.append(name)
            continue
        try:
            idx = max(0, min(int(it.get("frame_index", 0)), len(frames) - 1))
        except (TypeError, ValueError):
            idx = 0
        try:
            r = vlm.query_image_voted(frames[idx], prompt=prompt, model=model,
                                      mode=mode, api_key=api_key, passes=passes,
                                      reasoning=reasoning)
            verdicts.append({"name": name, "has_cat": has_cat, "answer": r["answer"],
                             "reason": r["reason"], "ratio": r["ratio"],
                             "unanimous": r["unanimous"], "borderline": r["borderline"],
                             "query_ms": r["query_ms"], "error": None})
        except RuntimeError as exc:     # a per-image query failure — keep going
            verdicts.append({"name": name, "has_cat": has_cat, "answer": None,
                             "reason": "", "ratio": "", "unanimous": False,
                             "borderline": True, "query_ms": 0.0, "error": str(exc)})
    return verdicts, skipped, cancelled


def _unmask_url(incoming: str, stored: str) -> str:
    """Keep ``stored`` when ``incoming`` is just its masked echo (the GUI sending
    back an unedited masked URL), else take ``incoming`` as a genuine edit.

    The GET responses mask a URL's inline password (``rtsp://user:pass@host`` →
    ``rtsp://user:***@host``, M1). Without this guard a routine re-save would
    persist that ``***`` mask over the real credential — the same trap the
    blank-password guards avoid for the dedicated password field.
    """
    incoming = incoming or ""
    stored = stored or ""
    if stored and incoming == mask_credentials(stored) and incoming != stored:
        return stored
    return incoming


def _mask_cameras(cameras, cfg=None) -> list:
    """Saved cameras with full per-camera settings, raw passwords stripped.

    Each entry is coerced to a complete camera dict (missing settings filled from
    the global defaults) so the GUI editor always has every field; the password is
    replaced by a ``has_password`` flag, and any inline URL credential is masked
    (M1: ``rtsp://user:pass@host`` must not leak via GET /api/cameras/saved).
    """
    out = []
    for c in cameras or []:
        if not isinstance(c, dict):
            continue
        full = config_mod.coerce_camera(c, cfg)
        full.pop("password", None)
        full["has_password"] = bool(c.get("password"))
        full["url"] = mask_credentials(str(full.get("url") or ""))
        out.append(full)
    return out


def _public_config(cfg) -> dict:
    """Config as a browser-safe dict: strip passwords, expand the camera list."""
    d = cfg.asdict()
    d.pop("camera_password", None)
    # Mask an inline URL credential in the legacy single-camera URL too (M1).
    d["camera_url"] = mask_credentials(str(d.get("camera_url") or ""))
    # The Moondream API key is a secret: never send it to the browser — expose only
    # whether one is set (the GUI shows "key set" / a paste field accordingly). #59
    d.pop("moondream_api_key", None)
    d["has_moondream_api_key"] = bool(cfg.moondream_api_key or os.environ.get("MOONDREAM_API_KEY"))
    d["cameras"] = _mask_cameras(d.get("cameras"), cfg)
    return d


# MJPEG stream cadence (review 2026-07-08, Finding 2). A client disconnect is only
# observable at a `yield`, so a stream that stops producing new frames must still
# yield periodically or its worker thread spins forever and leaks. HEARTBEAT re-emits
# the current frame on this cadence; NO_FRAME_TIMEOUT ends a stream that never got a
# frame at all (the browser retries on its next status tick). Module-level so tests
# can shrink them.
_STREAM_HEARTBEAT_S = 1.0
_STREAM_NO_FRAME_TIMEOUT_S = 10.0


def create_app(loop: DetectionLoop | None = None) -> Flask:
    app = Flask(__name__)
    app.config["loop"] = loop or DetectionLoop()
    app.config["MAX_CONTENT_LENGTH"] = _TEST_MAX_UPLOAD   # cap test-video uploads

    # -- page ---------------------------------------------------------------
    @app.get("/")
    def index():
        return send_from_directory(app.template_folder, "index.html")

    @app.get("/api/version")
    def api_version():
        return jsonify({"version": __version__})

    @app.get("/guide")
    def guide():
        # The user & workflow guide (docs/USER_GUIDE.md) served in-app so "what does
        # this button do?" has a one-click answer (the ❓ link in the header). One
        # source: the markdown in the repo IS the page — rendered with the tiny
        # converter below, no new dependency, no second copy to drift.
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "..", "docs", "USER_GUIDE.md")
        try:
            with open(path, encoding="utf-8") as fh:
                md = fh.read()
        except OSError:
            return Response("The guide file (docs/USER_GUIDE.md) is missing.",
                            mimetype="text/plain", status=404)
        return Response(_GUIDE_STYLE + _md_to_html(md), mimetype="text/html")

    @app.get("/api/models")
    def api_models():
        # The canonical [{value,label}] model list (present-checked), so every GUI
        # dropdown is built from the same registry the benchmark uses — no drift (#50).
        return jsonify({"models": _model_options()})

    # -- model provisioning (#86): audit the settled lineup, generate from GUI --
    @app.get("/api/models/audit")
    def api_models_audit():
        from . import provision, yolo

        return jsonify({"items": provision.audit(),
                        "can_provision": provision.ultralytics_available(),
                        "driver_cuda": yolo._driver_cuda_version(),
                        "running": _PROVISION["running"]})

    @app.post("/api/models/provision")
    def api_models_provision():
        from . import provision

        if _PROVISION["running"]:
            return jsonify({"error": "A provisioning run is already going."}), 409
        if not provision.ultralytics_available():
            return jsonify({"error": (
                "Provisioning needs the 'ultralytics' package on this machine "
                "(build-time only): pip install ultralytics — the app never "
                "installs it for you. Engines additionally need tensorrt + "
                "cuda-python on a CUDA-13 driver (#82).")}), 503
        body = request.get_json(silent=True) or {}
        _PROVISION.update({"running": True, "log": [], "error": None, "done": False})

        def run():
            try:
                provision.provision(targets=body.get("targets") or None,
                                    force=bool(body.get("force")),
                                    progress=_PROVISION["log"].append)
            except Exception as exc:      # noqa: BLE001 — surfaced to the GUI
                _PROVISION["error"] = str(exc)
            finally:
                _PROVISION["running"] = False
                _PROVISION["done"] = True

        threading.Thread(target=run, daemon=True).start()
        return jsonify({"started": True})

    @app.get("/api/models/provision/status")
    def api_models_provision_status():
        return jsonify(dict(_PROVISION))

    # -- detection snapshots (annotated images shown in the activity log) ----
    @app.get("/snapshots/<path:name>")
    def snapshot(name):
        directory = app.config["loop"].snapshots.directory
        if not os.path.exists(os.path.join(directory, name)):
            return jsonify({"error": "not found"}), 404
        return send_from_directory(directory, name)

    @app.get("/api/diagnostics/detectors")
    def api_diagnostics_detectors():
        """What each LIVE detector is actually holding, beside what you configured.

        Find builds its settings from the running detector's in-memory attributes
        (``det.model``, ``det.cat_confidence``, ``det.locator_classes`` …), not from
        the config file. So a hot-reload that didn't apply, or a ``reconfigure``
        that misses a field, makes find scan with settings you never set — while
        still reporting the values it was handed. This shows the difference without
        needing a cat to walk past.

        ``drift`` lists fields where the live detector disagrees with the saved
        camera spec. Anything in there is a bug.
        """
        loop = app.config["loop"]
        cfg = config_mod.load()
        specs = {s["name"]: s for s in config_mod.camera_targets(cfg)}
        # (live attribute, saved spec key) — the fields find actually reads.
        pairs = [("model", "model"), ("accelerator", "accelerator"),
                 ("confidence", "person_confidence"),
                 ("cat_confidence", "cat_confidence"),
                 ("label_floor", "label_floor"), ("roi", "roi"),
                 ("gamma", "gamma"), ("brightness", "brightness"),
                 ("contrast", "contrast"), ("saturation", "saturation"),
                 ("motion_hold_seconds", None)]
        out = []
        for name, det in (loop._detectors or {}).items():
            spec = specs.get(name) or {}
            live, saved, drift = {}, {}, []
            for attr, key in pairs:
                lv = getattr(det, attr, None)
                live[attr] = lv
                if key is None or key not in spec:
                    continue
                sv = spec[key]
                saved[attr] = sv
                if isinstance(lv, float) or isinstance(sv, float):
                    same = lv is not None and abs(float(lv) - float(sv)) < 1e-6
                else:
                    same = lv == sv
                if not same:
                    drift.append({"field": attr, "live": lv, "saved": sv})
            lc_live = list(getattr(det, "locator_classes", ()) or ())
            lc_saved = list(spec.get("locator_classes") or [])
            live["locator_classes"] = lc_live
            if lc_saved and lc_live != lc_saved:
                drift.append({"field": "locator_classes",
                              "live": lc_live, "saved": lc_saved})
            eff, why = det.effective_accelerator()
            out.append({"camera": name, "live": live, "saved": saved,
                        "ran_on": eff, "fallback": why,
                        "smooth": bool(getattr(det, "smooth_feed", False)),
                        "drift": drift})
        return jsonify({
            "detectors": out,
            "find": {"model": cfg.find_model or "(each camera's own)",
                     "tiling": cfg.find_tiling, "tile_overlap": cfg.find_tile_overlap,
                     "confidence": cfg.find_confidence or "(each camera's own)"},
            "any_drift": any(d["drift"] for d in out),
        })

    # -- manual screenshots from the live feed ------------------------------
    @app.get("/screenshots/<path:name>")
    def screenshot_file(name):
        directory = app.config["loop"].screenshots.directory
        if not os.path.exists(os.path.join(directory, name)):
            return jsonify({"error": "not found"}), 404
        return send_from_directory(directory, name)

    @app.post("/api/live/screenshot")
    def api_live_screenshot():
        """Save the camera's current picture, unannotated, to keep.

        Deliberately no boxes: this is for comparing a moment against the
        camera's own recordings later, and drawn overlays get in the way of that.
        Unlike detection snapshots these are never pruned — the user asked for
        this one, so the app doesn't get to decide it has expired.
        """
        loop = app.config["loop"]
        camera = ((request.get_json(silent=True) or {}).get("camera") or "").strip()
        if not loop.is_running():
            return jsonify({"error": "Start watching before taking a screenshot."}), 409
        det = loop.get_detector(camera) if camera else None
        if det is None:
            return jsonify({"error": f"Unknown camera {camera!r}."}), 409
        # PNG: these get re-run through the Test tool and compared against what
        # the live scan reported, so they have to be the pixels that were judged.
        img = det.plain_image(".png")
        if img is None:
            return jsonify({"error": "No frame from that camera yet."}), 409
        name = loop.screenshots.save(img, camera, ext=".png")
        if not name:
            return jsonify({"error": "Couldn't write the screenshot to disk."}), 500
        loop.activity.add("info", f"📸 Screenshot saved from {camera}: {name}")
        return jsonify({"ok": True, "file": name, "url": f"/screenshots/{name}"})

    # -- live preview frame (for the region-of-interest picker) -------------
    @app.get("/api/preview")
    def api_preview():
        cfg = config_mod.load()
        name = request.args.get("camera")
        if name:
            cam = next((c for c in (cfg.cameras or [])
                        if isinstance(c, dict) and c.get("name") == name), None)
            if cam is None:
                return jsonify({"error": "camera not found"}), 404
            source = config_mod.camera_source(
                cam.get("url", ""), cam.get("username", ""), cam.get("password", ""))
        elif cfg.camera_url:
            source = _camera_source(cfg)
        else:
            return jsonify({"error": "No camera configured yet."}), 400
        jpeg = grab_frame_jpeg(source)
        if jpeg is None:
            return jsonify({"error": "Couldn't grab a frame from the camera."}), 502
        return Response(jpeg, mimetype="image/jpeg")

    # -- live detection feed (MJPEG of what the running loop sees) -----------
    @app.get("/api/stream")
    def api_stream():
        loop = app.config["loop"]
        if not loop.is_running():
            return jsonify(
                {"error": "Start watching to see the live detection feed."}
            ), 409

        name = request.args.get("camera")     # which camera's feed (default: streamed one)
        # trail=1 (0.39.0): composite the cat trail as a LIVE overlay — the ribbon
        # and route line update in place as the cat moves (the 🌈 toggle by the feed).
        trail = request.args.get("trail") == "1"
        # last_known=0 (0.42.0) hides the default-on grey "last known location"
        # box (the camera's newest recorded sighting, age-labelled).
        last_known = request.args.get("last_known") != "0"

        def frames():
            # One JPEG per part; the browser renders this directly in an <img>.
            # Encode only when the frame/box version changes, so we never re-send
            # an unchanged frame and the feed runs at whatever rate frames arrive
            # — camera rate on a network camera, the loop's scan rate on USB.
            # The 0.03 s poll is the only ceiling (~30 fps); cheap when idle.
            #
            # A client disconnect is only raised (GeneratorExit) at a `yield`, so we
            # must keep yielding even when the frame version stalls: re-emit the last
            # frame every _STREAM_HEARTBEAT_S as a keepalive, and end a stream that
            # never produced a frame after _STREAM_NO_FRAME_TIMEOUT_S — otherwise this
            # generator (and its worker thread) spins forever and leaks (review
            # 2026-07-08, Finding 2). try/finally makes the exit path explicit.
            last_ver = -1
            last_jpeg = None
            last_emit = 0.0
            start = time.monotonic()
            try:
                while loop.is_running():
                    loop.note_viewing(name)   # keep this camera awake under round-robin
                    now = time.monotonic()
                    ver = loop.live_version(name)
                    out = None
                    if ver != last_ver:
                        jpeg = loop.live_jpeg(name, trail=trail, last_known=last_known)
                        if jpeg is not None:
                            last_ver, last_jpeg, out = ver, jpeg, jpeg
                    elif last_jpeg is not None and now - last_emit >= _STREAM_HEARTBEAT_S:
                        out = last_jpeg       # keepalive re-emit → disconnect noticed here
                    if out is not None:
                        last_emit = now
                        yield (b"--frame\r\nContent-Type: image/jpeg\r\n"
                               b"Content-Length: " + str(len(out)).encode() + b"\r\n\r\n"
                               + out + b"\r\n")
                    elif last_jpeg is None and now - start >= _STREAM_NO_FRAME_TIMEOUT_S:
                        return                # never got a frame; don't spin forever
                    time.sleep(0.03)
            finally:
                # Client disconnected or the loop stopped: nothing to release (the
                # generator's locals are GC'd), but the explicit exit path means a
                # stalled stream can no longer leak its worker thread.
                pass

        return Response(frames(),
                        mimetype="multipart/x-mixed-replace; boundary=frame")

    @app.get("/api/trail")
    def api_trail():
        # The cat trail (#67): the camera's current frame with the episode's
        # silhouettes tinted by recency (blue = entered → red = latest) and the
        # endpoint boxed. 404 when there's been no motion this episode.
        loop = app.config["loop"]
        if not loop.is_running():
            return jsonify({"error": "Start watching to see a cat trail."}), 409
        name = request.args.get("camera")
        det = loop.get_detector(name) if name else None
        if det is None:
            return jsonify({"error": f"Unknown camera {name!r}."}), 404
        raw = det.latest_frame()
        if raw is None:
            return jsonify({"error": "No frame from that camera yet."}), 409
        frame = det._crop(raw)
        jpeg = det.trail_jpeg(frame)
        if jpeg is None:
            return jsonify({"error": "No trail yet — nothing has moved this episode."}), 404
        return Response(jpeg, mimetype="image/jpeg")


    # -- discovery ----------------------------------------------------------
    @app.get("/api/speakers")
    def api_speakers():
        return jsonify(discovery.discover_speakers())

    @app.get("/api/cameras")
    def api_cameras():
        return jsonify(discovery.discover_cameras())

    @app.get("/api/cameras/local")
    def api_cameras_local():
        # USB/built-in cameras on the machine running the app.
        return jsonify(discovery.probe_local_cameras())

    # -- sounds -------------------------------------------------------------
    @app.get("/api/sounds")
    def api_sounds():
        files = sorted(
            f for f in os.listdir(config_mod.SOUNDS_DIR)
            if os.path.splitext(f)[1].lower() in ALLOWED_SOUND_EXT
        )
        return jsonify(files)

    @app.post("/api/sounds")
    def api_upload_sound():
        if "file" not in request.files:
            return jsonify({"error": "no file uploaded"}), 400
        f = request.files["file"]
        name = secure_filename(f.filename or "")
        if not name or os.path.splitext(name)[1].lower() not in ALLOWED_SOUND_EXT:
            return jsonify({"error": "unsupported file type"}), 400
        f.save(os.path.join(config_mod.SOUNDS_DIR, name))
        return jsonify({"saved": name})

    # -- config -------------------------------------------------------------
    @app.get("/api/config")
    def api_get_config():
        return jsonify(_public_config(config_mod.load()))

    @app.post("/api/config")
    def api_set_config():
        values = request.get_json(silent=True) or {}
        # Don't overwrite a stored password with an empty form field.
        if not values.get("camera_password"):
            values.pop("camera_password", None)
        # GET masks an inline URL credential; don't let its masked echo overwrite
        # the real stored URL on a routine save (M1).
        if "camera_url" in values:
            values["camera_url"] = _unmask_url(values.get("camera_url"),
                                               config_mod.load().camera_url)
        # Same for the Moondream API key: a blank field on save keeps the stored key
        # (the GUI only sends it when the user actually types a new one). #59
        if not (values.get("moondream_api_key") or "").strip():
            values.pop("moondream_api_key", None)
        # The saved-camera store is managed only via the /api/cameras/saved
        # endpoints, so the main settings save can't clobber it (or its passwords).
        values.pop("cameras", None)
        cfg = config_mod.update(values)
        return jsonify(_public_config(cfg))

    # -- saved cameras (full per-camera config + credentials) ----------------
    @app.get("/api/cameras/saved")
    def api_cameras_saved():
        cfg = config_mod.load()
        return jsonify(_mask_cameras(cfg.cameras, cfg))

    @app.post("/api/cameras/saved")
    def api_cameras_save():
        data = request.get_json(silent=True) or {}
        name = (data.get("name") or "").strip()
        url = (data.get("url") or "").strip()
        if not name or not url:
            return jsonify({"error": "A camera needs a name and a stream URL."}), 400
        cfg = config_mod.load()
        cams = [c for c in (cfg.cameras or []) if isinstance(c, dict)]
        existing = next((c for c in cams if c.get("name") == name), None)
        # GET masks an inline URL credential; if the client echoed that masked URL
        # back unchanged, keep the stored one instead of persisting the mask (M1).
        if existing is not None:
            url = _unmask_url(url, existing.get("url", ""))
        # Start from the existing entry (so unspecified settings persist), overlay
        # the incoming fields, then coerce to a complete per-camera dict.
        merged = dict(existing or {})
        merged.update(data)
        merged["name"], merged["url"] = name, url
        entry = config_mod.coerce_camera(merged, cfg)
        # A blank password on re-save keeps the previously-stored one.
        if not (data.get("password") or "").strip() and existing is not None:
            entry["password"] = existing.get("password", "")
        if existing is not None:
            cams[cams.index(existing)] = entry
        else:
            cams.append(entry)
        config_mod.update({"cameras": cams})
        return jsonify(_mask_cameras(cams, cfg))

    @app.post("/api/cameras/saved/select")
    def api_cameras_select():
        name = ((request.get_json(silent=True) or {}).get("name") or "").strip()
        cfg = config_mod.load()
        cam = next((c for c in (cfg.cameras or [])
                    if isinstance(c, dict) and c.get("name") == name), None)
        if cam is None:
            return jsonify({"error": "camera not found"}), 404
        config_mod.update({
            "camera_name": cam.get("name", ""),
            "camera_url": cam.get("url", ""),
            "camera_username": cam.get("username", ""),
            "camera_password": cam.get("password", ""),
        })
        return jsonify(_public_config(config_mod.load()))

    @app.post("/api/cameras/saved/delete")
    def api_cameras_delete():
        name = ((request.get_json(silent=True) or {}).get("name") or "").strip()
        cfg = config_mod.load()
        cams = [c for c in (cfg.cameras or [])
                if isinstance(c, dict) and c.get("name") != name]
        # Also drop it from the watched set if present.
        active = [n for n in (cfg.active_cameras or []) if n != name]
        config_mod.update({"cameras": cams, "active_cameras": active})
        return jsonify(_mask_cameras(cams, cfg))

    @app.post("/api/cameras/active")
    def api_cameras_active():
        """Set which saved cameras are watched simultaneously (multi-camera)."""
        names = (request.get_json(silent=True) or {}).get("names") or []
        cfg = config_mod.load()
        valid = {c.get("name") for c in (cfg.cameras or []) if isinstance(c, dict)}
        active = [n for n in names if n in valid]
        config_mod.update({"active_cameras": active})
        return jsonify({"active_cameras": active})

    # -- control ------------------------------------------------------------
    @app.post("/api/test")
    def api_test():
        try:
            app.config["loop"].test_cast()
            return jsonify({"ok": True})
        except Exception as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400

    @app.post("/api/start")
    def api_start():
        started = app.config["loop"].start()
        return jsonify({"running": True, "started": started})

    @app.post("/api/stop")
    def api_stop():
        app.config["loop"].stop()
        return jsonify({"running": False})

    @app.get("/api/status")
    def api_status():
        loop = app.config["loop"]
        s = loop.status
        return jsonify(
            {
                "running": loop.is_running(),
                "last_error": s.last_error,
                "last_roll": s.last_roll,
                "last_roll_at": s.last_roll_at,
                "rolls": s.rolls,
                "treats": s.treats,
                "cameras": loop.cam_status(),   # per-camera connected/error + roles
            }
        )

    # -- activity log -------------------------------------------------------
    @app.get("/api/log")
    def api_log():
        limit = request.args.get("limit", default=200, type=int)
        return jsonify(app.config["loop"].activity.entries(limit=limit))

    @app.post("/api/log/clear")
    def api_log_clear():
        app.config["loop"].activity.clear()
        return jsonify({"ok": True})

    # -- cat sightings ("show cat") -----------------------------------------
    @app.get("/api/cats")
    def api_cats():
        limit = request.args.get("limit", default=20, type=int)
        loop = app.config["loop"]
        return jsonify({
            "last": loop.cats.last(),
            "today": loop.cats.today_count(),
            "present": loop.cat_present(),     # cat on camera right now → flash the button
            "cameras": loop.cats_present_cameras(),   # cameras seeing a cat now → Show-cat rotation
            "recent": loop.cats.recent(limit=limit),
            # Time-of-day PRIOR (#68): where the cats usually are around this hour —
            # ranks a Find-My-Cat sweep; a hint from history, never a tracked state.
            "by_hour": loop.cats.by_hour(),
            "button": loop.button_state(),     # the Cat-cam button's ambient glyph
            "likely": [{"camera": c, "weight": n} for c, n in loop.cats.likely_cameras()],
        })

    @app.get("/api/cats/heatmap")
    def api_cats_heatmap():
        # Sighting-density heat map over the camera's current frame (#68) — the
        # camera's hot spots (basket / couch arm / sunny patch) from cats.log.
        from . import heatmap as heatmap_mod

        loop = app.config["loop"]
        if not loop.is_running():
            return jsonify({"error": "Start watching to see a heat map."}), 409
        name = request.args.get("camera")
        det = loop.get_detector(name) if name else None
        if det is None:
            return jsonify({"error": f"Unknown camera {name!r}."}), 404
        raw = det.latest_frame()
        if raw is None:
            return jsonify({"error": "No frame from that camera yet."}), 409
        frame = det._crop(raw)
        boxes = [s["box"] for s in loop.cats.recent() if s.get("camera") == name]
        img = heatmap_mod.render_heatmap(frame, boxes)
        if img is None:
            return jsonify({"error": "No sightings recorded for this camera yet."}), 404
        import cv2
        ok, buf = cv2.imencode(".jpg", img, [cv2.IMWRITE_JPEG_QUALITY, 80])
        if not ok:
            return jsonify({"error": "Couldn't encode the heat map."}), 500
        return Response(buf.tobytes(), mimetype="image/jpeg")

    @app.post("/api/cats/find")
    def api_cats_find():
        # "Show me the cat" ACTIVE scan (#92): a real detection pass across the
        # find cameras on click, so a still cat in a motionless room is found —
        # search, don't just jump to the last sighting. Runs the tester's cached
        # per-(model, accelerator) nets on each camera's LATEST frame — never the
        # worker's net, so there's no race with the live loop. Finds are recorded
        # (source "find") and the best camera is boosted so the feed shows the box.
        import cv2      # for encoding the scanned frame, and via detect_image

        loop = app.config["loop"]
        cfg = config_mod.load()
        if not cfg.find_scan:
            return jsonify({"error": "Active find-scan is off — enable it in the "
                            "Find-my-cat settings (the button then jumps to the "
                            "last sighting as before)."}), 403
        if not loop.is_running():
            return jsonify({"error": "Detection isn't running."}), 409
        wanted = {c for c in (cfg.find_cameras or []) if c}
        # Only scan WATCHED cameras (#103): the live active-cameras set, not
        # everything the loop happens to still hold. An empty active set means
        # the single legacy camera (camera_targets handles that).
        watched = {s.get("name") for s in config_mod.camera_targets(cfg)}
        zones_by_cam = {c.get("name"): c.get("zones") or []
                        for c in (cfg.cameras or []) if isinstance(c, dict)}
        results = []
        for name, det in loop._detectors.items():
            if name not in watched:
                continue
            if wanted and name not in wanted:
                continue
            raw = det.latest_frame()
            if raw is None:
                results.append({"camera": name, "found": False, "score": 0.0,
                                "note": "no frame yet"})
                continue
            frame = det._crop(raw)
            # Find has its own full settings (#101): model/tiling/overlap/
            # confidence, each with "each camera's own" (empty/0) as the default.
            settings = {
                "model": cfg.find_model or det.model,
                "accelerator": det.accelerator,
                "person_confidence": det.confidence,
                "cat_confidence": cfg.find_confidence or det.cat_confidence,
                "label_floor": det.label_floor,
                "locator_classes": list(det.locator_classes),
                "tiling": cfg.find_tiling or "3x3",
                "tile_overlap": cfg.find_tile_overlap,
            }
            ran = {}
            annotated, dets, ms = _run_test_detection(frame, settings, report=ran)
            best = max((d for d in dets if d["label"] in det.locator_classes),
                       key=lambda d: d["score"], default=None)
            found = bool(best and best["score"]
                         >= (cfg.find_confidence or det.cat_confidence))
            row = {"camera": name, "found": found, "ms": ms,
                   "score": round(best["score"], 3) if best else 0.0}
            # Keep BOTH images from every find, found or not: the exact frame that
            # was scanned, and what the scan made of it. A find that reports
            # nothing is otherwise unfalsifiable — you can only re-take a
            # screenshot afterwards and hope it's the same moment, which it isn't.
            # Also record what it actually ran with, so "no cat found" is a
            # checkable statement rather than one you have to take on trust.
            row["asked"] = dict(settings)     # what find requested…
            row["ran"] = ran                  # …and what the detector actually used
            row["everything_seen"] = [
                {"label": d["label"], "score": d["score"]} for d in dets[:8]
            ]
            # The best locator-class box REGARDLESS of the threshold. Without it,
            # "no cat" can't distinguish "the net never saw one" from "it saw one
            # and we rejected it" — and those need completely different fixes.
            row["best_locator"] = ({"label": best["label"], "score": best["score"]}
                                   if best else None)
            # A find is a sweep of that room, so its verdict replaces the room's
            # previous one — same as a periodic still-scan.
            loop.record_sweep(name, found)
            stamp = loop.find_shots.stamp()
            # PNG, not JPEG: the scanned frame is evidence, and it has to be the
            # pixels that were actually judged. A borderline detection sits close
            # enough to its threshold that JPEG requantisation alone can flip it,
            # so a lossy copy replays as a DIFFERENT scan and quietly disagrees
            # with the run it is supposed to explain. Costs disk; worth it.
            ok_in, buf_in = cv2.imencode(".png", frame, [cv2.IMWRITE_PNG_COMPRESSION, 3])
            row["scanned_image"] = loop.find_shots.save(
                buf_in.tobytes() if ok_in else None, name, "scanned", stamp, ext=".png")
            row["result_image"] = loop.find_shots.save(annotated, name, "result", stamp)
            if found:
                snap = loop.snapshots.save(annotated)
                h, w = frame.shape[:2]
                sighting = loop.cats.record(
                    name, best["box"], (w, h), best["score"], image=snap,
                    label=best["label"], source="find",
                    zone=zone_for(best["box"], zones_by_cam.get(name), det.roi))
                spot = sighting.get("zone") or sighting["region"]
                row["where"] = spot or ""
                loop.activity.add(
                    "motion",
                    f"🔎 Find-scan: cat found{f' ({spot})' if spot else ''} on "
                    f"{name}.", image=snap)
            results.append(row)
        hits = [r for r in results if r["found"]]
        best_cam = max(hits, key=lambda r: r["score"], default=None)
        if best_cam:
            loop.boost_detection(best_cam["camera"], 10.0)
        # One log line naming where the evidence went, so a find that reports
        # nothing can be re-examined later instead of argued about.
        saved = [r["scanned_image"] for r in results if r.get("scanned_image")]
        if saved:
            loop.activity.add(
                "info",
                f"🔎 Find-scan kept {len(saved) * 2} images "
                f"(screenshots/find/) — the frames scanned and what each scan saw.")
        return jsonify({"results": results,
                        "found": [r["camera"] for r in hits],
                        "images_dir": "screenshots/find",
                        "best": best_cam["camera"] if best_cam else None})

    @app.get("/api/feeds")
    def api_feeds():
        """Which camera each live feed should show (#113, Follow mode).

        Sticky/debounced assignment — see :mod:`d20app.feeds`. ``slots`` is how
        many feeds the GUI is showing (1 or 2).
        """
        try:
            slots = max(1, min(2, int(request.args.get("slots", 1))))
        except (TypeError, ValueError):
            slots = 1
        # Locks are per-feed and live in the browser (a runtime toggle, not saved),
        # so they arrive with the request: ?locked=0,1
        locks = set()
        for part in (request.args.get("locked") or "").split(","):
            if part.strip().isdigit():
                locks.add(int(part.strip()))
        cfg = config_mod.load()
        return jsonify({"slots": app.config["loop"].feed_assignments(
            slots, locks=locks, confirm=cfg.swap_confirm_count,
            reuse_cooldown=cfg.camera_reuse_cooldown_seconds)})

    @app.post("/api/cats/clear")
    def api_cats_clear():
        loop = app.config["loop"]
        loop.cats.clear()
        loop.clear_last_known()   # #112: also drop the live overlay box, not just the log
        return jsonify({"ok": True})

    # -- "Test detection" tool (upload a photo/video, tune settings, draw boxes) --
    @app.post("/api/test/upload")
    def api_test_upload():
        if "file" not in request.files:
            return jsonify({"error": "no file uploaded"}), 400
        data = request.files["file"].read()
        if not data:
            return jsonify({"error": "empty file"}), 400
        # Optional frame count for videos (#38) — default ~1 fps when omitted.
        try:
            n_frames = int(request.form.get("frames")) if request.form.get("frames") else None
        except (TypeError, ValueError):
            n_frames = None
        frames = _decode_test_upload(data, request.files["file"].filename or "", n_frames)
        if not frames:
            return jsonify({"error": "Couldn't read that as an image or video."}), 400
        sid = uuid.uuid4().hex
        with _TEST_SESSIONS_LOCK:
            _TEST_SESSIONS[sid] = frames
            while len(_TEST_SESSIONS) > _TEST_MAX_SESSIONS:
                _TEST_SESSIONS.popitem(last=False)     # drop the oldest session
        h, w = frames[0].shape[:2]
        return jsonify({
            "id": sid, "count": len(frames), "width": int(w), "height": int(h),
            "kind": "video" if len(frames) > 1 else "image",
            "thumbs": [_thumb_data_url(fr) for fr in frames],
        })

    @app.post("/api/test/detect")
    def api_test_detect():
        body = request.get_json(silent=True) or {}
        with _TEST_SESSIONS_LOCK:
            frames = _TEST_SESSIONS.get(body.get("id"))
        if not frames:
            return jsonify({"error": "Upload expired — pick the file again."}), 404
        try:
            idx = int(body.get("frame_index", 0))
        except (TypeError, ValueError):
            idx = 0
        idx = max(0, min(idx, len(frames) - 1))
        annotated, dets, ms = _run_test_detection(frames[idx], body.get("settings") or {})
        if annotated is None:
            return jsonify({"error": "Detection failed on that frame."}), 500
        return jsonify({
            "annotated": "data:image/jpeg;base64," + base64.b64encode(annotated).decode("ascii"),
            "detections": dets, "frame_index": idx, "inference_ms": ms,
        })

    @app.get("/api/test/benchmark/models")
    def api_benchmark_models():
        return jsonify({"models": _benchmark_models(), "tilings": _BENCH_TILINGS})

    @app.post("/api/test/benchmark")
    def api_test_benchmark():
        body = request.get_json(silent=True) or {}
        with _TEST_SESSIONS_LOCK:
            frames = _TEST_SESSIONS.get(body.get("id"))
        if not frames:
            return jsonify({"error": "Upload expired — pick the file again."}), 404
        try:
            idx = max(0, min(int(body.get("frame_index", 0)), len(frames) - 1))
        except (TypeError, ValueError):
            idx = 0
        models = [m for m in (body.get("models") or _benchmark_models()) if m]
        tilings = [t for t in (body.get("tilings") or _BENCH_TILINGS) if t]
        if not models or not tilings:
            return jsonify({"error": "Pick at least one model and tiling."}), 400
        if len(models) * len(tilings) > _BENCH_MAX_RUNS:
            return jsonify({"error": f"Too many runs ({len(models) * len(tilings)}); "
                            f"trim to ≤ {_BENCH_MAX_RUNS}."}), 400
        try:
            cat_threshold = float(body.get("cat_threshold",
                                           config_mod.load().cat_confidence))
        except (TypeError, ValueError):
            cat_threshold = 0.5
        accelerator = body.get("accelerator") or "cpu"
        opts = _bench_opts(body)        # tester controls, applied to every run (#44)
        runs = _run_benchmark(frames[idx], models, tilings, cat_threshold, accelerator, opts)
        h, w = frames[idx].shape[:2]
        image_name = body.get("name") or "uploaded image"
        meta = {"cat_threshold": cat_threshold, "accelerator": accelerator,
                "image": image_name,
                "ts": time.strftime("%Y-%m-%d %H:%M"), "width": int(w), "height": int(h),
                **opts}
        # A human-readable filename slug for the downloaded report (#27), and the
        # full-res original frame embedded once for reproducibility (#25).
        stem = os.path.splitext(os.path.basename(image_name))[0]
        slug = f"benchmark-{_slugify(stem)}-{time.strftime('%Y%m%d-%H%M%S')}"
        original_url = _full_jpeg_data_url(frames[idx])
        original_name = f"{_slugify(stem)}.jpg"
        html = _benchmark_html(runs, meta, original_url=original_url,
                               original_name=original_name)
        try:
            xlsx = _benchmark_xlsx(runs, meta)
            xlsx_error = None
        except Exception as exc:        # noqa: BLE001 — optional dep / report failure
            xlsx, xlsx_error = None, str(exc)
        rid = uuid.uuid4().hex
        with _BENCH_LOCK:
            _BENCHMARKS[rid] = {"html": html, "xlsx": xlsx, "slug": slug}
            while len(_BENCHMARKS) > _BENCH_MAX_REPORTS:
                _BENCHMARKS.popitem(last=False)
        public = [{k: v for k, v in r.items() if k != "_raw"} for r in runs]
        return jsonify({
            "id": rid, "runs": public, "meta": meta,
            "html_url": f"/api/test/benchmark/{rid}.html",
            "xlsx_url": (f"/api/test/benchmark/{rid}.xlsx" if xlsx else None),
            "xlsx_error": xlsx_error,
        })

    @app.post("/api/test/benchmark/batch")
    def api_test_benchmark_batch():
        # Run the sweep across many uploaded images, emit a per-image report for each
        # plus one cross-image summary that aggregates which config is most reliable.
        body = request.get_json(silent=True) or {}
        items = body.get("items") or []
        if not items:
            return jsonify({"error": "Add at least one image to the batch."}), 400
        # Soft cap (#37): the recommended limit is advisory — the GUI warns with the
        # cost and proceeds. Only an absolute ceiling is refused, to bound pathology.
        if len(items) > _BENCH_MAX_IMAGES_HARD:
            return jsonify({"error": f"Too many images ({len(items)}); the absolute "
                            f"ceiling is {_BENCH_MAX_IMAGES_HARD}."}), 400
        models = [m for m in (body.get("models") or _benchmark_models()) if m]
        tilings = [t for t in (body.get("tilings") or _BENCH_TILINGS) if t]
        if not models or not tilings:
            return jsonify({"error": "Pick at least one model and tiling."}), 400
        if len(models) * len(tilings) > _BENCH_MAX_RUNS:
            return jsonify({"error": f"Too many runs per image "
                            f"({len(models) * len(tilings)}); trim to ≤ {_BENCH_MAX_RUNS}."}), 400
        try:
            cat_threshold = float(body.get("cat_threshold",
                                           config_mod.load().cat_confidence))
        except (TypeError, ValueError):
            cat_threshold = 0.5
        accelerator = body.get("accelerator") or "cpu"
        opts = _bench_opts(body)               # tester controls, applied to every run (#44)
        batch_id = body.get("batch_id")        # client token for mid-run abort (#37)
        ts_disp = time.strftime("%Y-%m-%d %H:%M")
        ts_file = time.strftime("%Y%m%d-%H%M%S")
        # Optional "also run VLM on each image" (#54). Off by default — no added cost.
        # If the VLM can't run, don't fail the whole sweep: note it and carry on.
        run_vlm = bool(body.get("run_vlm"))
        vlm_model = body.get("vlm_model") or vlm.DEFAULT_MODEL
        vlm_mode = body.get("vlm_mode") or vlm.DEFAULT_MODE
        vlm_passes = body.get("vlm_passes") or vlm.DEFAULT_PASSES
        vlm_key = _vlm_key(body)
        vlm_error = None
        if run_vlm:
            try:
                vlm.preflight(vlm_key, vlm_mode)
            except RuntimeError as exc:
                run_vlm, vlm_error = False, str(exc)

        per_image, images, members, skipped = [], [], [], []
        cancelled = False
        for i, it in enumerate(items):
            # Abort check between images: stop, keep what's done, return partial (#37).
            if batch_id:
                with _BENCH_CANCEL_LOCK:
                    if batch_id in _BENCH_CANCEL:
                        cancelled = True
            if cancelled:
                break
            name = (it or {}).get("name") or f"image {i + 1}"
            with _TEST_SESSIONS_LOCK:
                frames = _TEST_SESSIONS.get((it or {}).get("id"))
            if not frames:
                # Don't hide it: a missing session means a quietly-incomplete run (#40).
                skipped.append(name)
                continue
            try:
                idx = max(0, min(int(it.get("frame_index", 0)), len(frames) - 1))
            except (TypeError, ValueError):
                idx = 0
            frame = frames[idx]
            has_cat = it.get("has_cat", True) is not False
            runs = _run_benchmark(frame, models, tilings, cat_threshold, accelerator, opts)
            h, w = frame.shape[:2]
            stem = os.path.splitext(os.path.basename(name))[0]
            meta_i = {"cat_threshold": cat_threshold, "accelerator": accelerator,
                      "image": name, "ts": ts_disp, "width": int(w), "height": int(h),
                      **opts}
            rid = uuid.uuid4().hex
            slug = f"benchmark-{_slugify(stem)}-{ts_file}-{i + 1}"
            html = _benchmark_html(runs, meta_i,
                                   original_url=_full_jpeg_data_url(frame),
                                   original_name=f"{_slugify(stem)}.jpg")
            with _BENCH_LOCK:
                _BENCHMARKS[rid] = {"html": html, "xlsx": None, "slug": slug}
            members.append((slug, html))
            # Keep only the light fields the summary needs (no thumbs/raw bytes).
            light = [{"model": r["model"], "tiling": r["tiling"], "size": r["size"],
                      "combined_score": r["combined_score"], "detected": r["detected"],
                      "inference_ms": r["inference_ms"]} for r in runs]
            entry = {"idx": len(per_image), "name": name, "report_id": rid,
                     "slug": slug, "has_cat": has_cat,
                     "orig_thumb": _orig_thumb_data_url(frame), "runs": light}
            if run_vlm:        # majority-voted VLM query per image, alongside the sweep (#54/#60)
                try:
                    vr = vlm.query_image_voted(frame, model=vlm_model, mode=vlm_mode,
                                               api_key=vlm_key, passes=vlm_passes)
                    entry["vlm"] = {"answer": vr["answer"], "reason": vr["reason"],
                                    "ratio": vr["ratio"], "unanimous": vr["unanimous"],
                                    "borderline": vr["borderline"],
                                    "query_ms": vr["query_ms"], "error": None}
                except RuntimeError as exc:
                    entry["vlm"] = {"answer": None, "reason": "", "ratio": "",
                                    "unanimous": False, "borderline": True,
                                    "query_ms": 0.0, "error": str(exc)}
            per_image.append(entry)
            # Link by slug so the link matches the per-image download filename (#35).
            images.append({"name": name, "has_cat": has_cat, "slug": slug,
                           "report_url": f"/api/test/benchmark/{slug}.html"})

        if batch_id:        # done (or aborted) — clear the cancel flag either way
            with _BENCH_CANCEL_LOCK:
                _BENCH_CANCEL.discard(batch_id)
        if not per_image:
            msg = ("Batch cancelled before any image finished." if cancelled
                   else "All uploads expired — re-select the images.")
            return jsonify({"error": msg, "cancelled": cancelled}), 404

        configs = _aggregate_configs(per_image)
        smeta = {"cat_threshold": cat_threshold, "accelerator": accelerator,
                 "ts": ts_disp, **opts}
        # VLM verdicts (if the toggle ran): recall/FP summary + the interesting frames
        # where the VLM and the best YOLO config disagree (#54).
        vlm_info = None
        graded = [im for im in per_image if im.get("vlm")]
        if graded:
            verdicts = [{**im["vlm"], "has_cat": im["has_cat"], "name": im["name"]}
                        for im in graded]
            disagreements = []
            for im in graded:
                v = im["vlm"]
                if v.get("error") or v.get("answer") is None:
                    continue
                yolo_found = any(r["detected"] for r in im["runs"])
                vlm_yes = v["answer"] == "yes"
                if vlm_yes != yolo_found:
                    disagreements.append(
                        {"name": im["name"], "slug": im["slug"], "has_cat": im["has_cat"],
                         "vlm": "yes" if vlm_yes else "no", "ratio": v.get("ratio", ""),
                         "yolo": "found" if yolo_found else "miss"})
            # Borderline (non-unanimous) frames — the genuinely-hard ones worth review (#60).
            borderline = [{"name": im["name"], "slug": im["slug"],
                           "answer": im["vlm"].get("answer"), "ratio": im["vlm"].get("ratio", "")}
                          for im in graded
                          if not im["vlm"].get("error") and im["vlm"].get("borderline")]
            vlm_info = {"summary": _vlm_summary(verdicts), "model": vlm_model,
                        "passes": vlm_passes, "disagreements": disagreements,
                        "borderline": borderline}
        elif vlm_error:
            vlm_info = {"error": vlm_error}
        summary_html = _benchmark_summary_html(per_image, configs, smeta, vlm=vlm_info)
        sid = uuid.uuid4().hex
        summary_slug = f"benchmark-summary-{ts_file}"
        with _BENCH_LOCK:
            # Stash the batch's members (slug → html) on the summary so "Download all"
            # can zip the summary + every per-image report under their slug names (#35).
            _BENCHMARKS[sid] = {"html": summary_html, "xlsx": None, "slug": summary_slug,
                                "members": [(summary_slug, summary_html)] + members}
            while len(_BENCHMARKS) > _BENCH_MAX_REPORTS:
                _BENCHMARKS.popitem(last=False)
        n_cat = sum(1 for im in per_image if im["has_cat"])
        return jsonify({
            "summary_id": sid,
            "summary_url": f"/api/test/benchmark/{sid}.html",
            "zip_url": f"/api/test/benchmark/{sid}/all.zip",
            "cancelled": cancelled,
            # Surface dropped images so a partial run can't masquerade as complete (#40).
            "requested": len(items), "ran": len(per_image),
            "skipped": len(skipped), "skipped_names": skipped[:20],
            "images": images, "configs": configs, "vlm": vlm_info,
            "meta": {"n_images": len(per_image), "n_cat": n_cat,
                     "n_nocat": len(per_image) - n_cat,
                     "cat_threshold": cat_threshold, "accelerator": accelerator},
        })

    @app.post("/api/test/benchmark/cancel")
    def api_test_benchmark_cancel():
        # Flag a running batch to stop after its current image (#37). The in-flight
        # batch request sees this between images and returns whatever finished.
        bid = (request.get_json(silent=True) or {}).get("batch_id")
        if not bid:
            return jsonify({"error": "no batch_id"}), 400
        with _BENCH_CANCEL_LOCK:
            _BENCH_CANCEL.add(bid)
        return jsonify({"ok": True})

    # -- local VLM (moondream) cat-presence tester (#48) ----------------------
    @app.get("/api/vlm/status")
    def api_vlm_status():
        # Tells the GUI whether the optional 'moondream' package is installed and whether
        # an API key is set, plus the coherent (mode, model) choices it should offer (#59),
        # so it can hint how to enable the tester instead of failing on Run.
        cfg = config_mod.load()
        return jsonify({"available": vlm.is_available(),
                        "has_api_key": bool(cfg.moondream_api_key
                                            or os.environ.get("MOONDREAM_API_KEY")),
                        "default_prompt": vlm.DEFAULT_PROMPT,
                        # Prompts are model-specific (#74): P6 on M3 → ~100% FP. The
                        # GUI swaps defaults on model change and warns on a carried
                        # custom prompt.
                        "model_prompts": vlm.MODEL_PROMPTS,
                        "choices": vlm.VLM_CHOICES, "default_choice": vlm.DEFAULT_CHOICE,
                        "models": list(vlm.MODELS), "default_model": vlm.DEFAULT_MODEL,
                        "default_passes": vlm.DEFAULT_PASSES, "max_passes": vlm.MAX_PASSES,
                        "escalation": {"enabled": bool(cfg.vlm_escalation)}})

    def _vlm_key(body) -> str | None:
        # Per-request key wins (e.g. a paste before saving); else the stored config key;
        # else the module falls back to MOONDREAM_API_KEY. Never logged. #59
        return (body.get("api_key") or "").strip() or config_mod.load().moondream_api_key or None

    @app.post("/api/vlm/query")
    def api_vlm_query():
        # One moondream 'query' pass on an uploaded/extracted frame. Runs on this
        # request's worker thread (Flask is threaded) — a slow query blocks only this
        # request, not the app; the GUI shows a "working…" state. Degrades with a clear
        # message when moondream, a GPU/VRAM, or the key is absent (never a 500).
        body = request.get_json(silent=True) or {}
        with _TEST_SESSIONS_LOCK:
            frames = _TEST_SESSIONS.get(body.get("id"))
        if not frames:
            return jsonify({"error": "Upload a frame first (it may have expired)."}), 404
        try:
            idx = max(0, min(int(body.get("frame_index", 0)), len(frames) - 1))
        except (TypeError, ValueError):
            idx = 0
        model = body.get("model") or vlm.DEFAULT_MODEL
        try:
            result = vlm.query_image_voted(
                frames[idx], prompt=body.get("prompt") or vlm.default_prompt(model),
                model=model,
                mode=body.get("mode") or vlm.DEFAULT_MODE, api_key=_vlm_key(body),
                passes=body.get("passes") or vlm.DEFAULT_PASSES,
                reasoning=bool(body.get("reasoning")))      # M3's thinking mode (#76)
        except RuntimeError as exc:         # optional dep / no GPU / no key / query failure
            return jsonify({"error": str(exc)}), 503
        return jsonify(result)

    @app.post("/api/vlm/locate")
    def api_vlm_locate():
        # moondream 'detect' mode exposed in the tester (#75): "show me WHERE the
        # model thinks a cat is on this frame". The diagnostic use case is a
        # false-positive null — the returned region says what feature (a cushion, a
        # shadow) triggers the phantom, which informs better exclusion prompts.
        # Purely diagnostic: nothing here records or confirms anything (#69 rules).
        import cv2

        body = request.get_json(silent=True) or {}
        with _TEST_SESSIONS_LOCK:
            frames = _TEST_SESSIONS.get(body.get("id"))
        if not frames:
            return jsonify({"error": "Upload a frame first (it may have expired)."}), 404
        try:
            idx = max(0, min(int(body.get("frame_index", 0)), len(frames) - 1))
        except (TypeError, ValueError):
            idx = 0
        frame = frames[idx]
        obj = (body.get("object") or "cat").strip() or "cat"
        mode = body.get("mode") or vlm.DEFAULT_MODE
        try:
            vlm.preflight(_vlm_key(body), mode)
            det = vlm.detect_regions(frame, obj,
                                     model=body.get("model") or vlm.DEFAULT_MODEL,
                                     api_key=_vlm_key(body), mode=mode)
        except RuntimeError as exc:
            return jsonify({"error": str(exc)}), 503
        h, w = frame.shape[:2]
        boxes = [b for b in (escalation.map_normalized_box(r, (w, h))
                             for r in det.get("objects") or []) if b]
        annotated = frame.copy()
        for i, (x1, y1, x2, y2) in enumerate(boxes):
            cv2.rectangle(annotated, (x1, y1), (x2, y2), (60, 120, 255), 2)
            cv2.putText(annotated, f"{obj} {i + 1}", (x1, max(14, y1 - 6)),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.55, (60, 120, 255), 2, cv2.LINE_AA)
        return jsonify({
            "object": obj, "n": len(boxes), "boxes": [list(b) for b in boxes],
            "detect_ms": det.get("detect_ms"), "model": det.get("model"),
            "mode": det.get("mode"), "device": det.get("device"),
            "annotated": _full_jpeg_data_url(annotated, quality=80),
            "note": ("" if boxes else
                     f"The model proposed no {obj} regions on this frame.")})

    @app.post("/api/vlm/batch")
    def api_vlm_batch():
        # Run the VLM once per uploaded image and score recall (cat-present frames) +
        # false-positive rate (no-cat frames) separately, against the same "cat present"
        # toggle the detection benchmark uses (#54). One query per image — no sweep.
        body = request.get_json(silent=True) or {}
        items = body.get("items") or []
        if not items:
            return jsonify({"error": "Add at least one image."}), 400
        if len(items) > _BENCH_MAX_IMAGES_HARD:
            return jsonify({"error": f"Too many images ({len(items)}); ceiling is "
                            f"{_BENCH_MAX_IMAGES_HARD}."}), 400
        model = body.get("model") or vlm.DEFAULT_MODEL
        mode = body.get("mode") or vlm.DEFAULT_MODE
        key = _vlm_key(body)
        try:                                # fail fast on a setup error, not per image
            vlm.preflight(key, mode)
        except RuntimeError as exc:
            return jsonify({"error": str(exc)}), 503
        batch_id = body.get("batch_id")
        passes = body.get("passes") or vlm.DEFAULT_PASSES
        verdicts, skipped, cancelled = _run_vlm_items(
            items, model, batch_id, mode, key, passes,
            prompt=body.get("prompt"),                      # the user's prompt (#72)
            reasoning=bool(body.get("reasoning")))
        if batch_id:
            with _BENCH_CANCEL_LOCK:
                _BENCH_CANCEL.discard(batch_id)
        if not verdicts:
            return jsonify({"error": "All uploads expired — re-select the images.",
                            "cancelled": cancelled}), 404
        return jsonify({
            "model": model, "passes": passes, "summary": _vlm_summary(verdicts),
            "verdicts": verdicts, "cancelled": cancelled, "requested": len(items),
            "ran": len(verdicts), "skipped": len(skipped), "skipped_names": skipped[:20],
        })

    @app.post("/api/vlm/escalate")
    def api_vlm_escalate():
        # The "find the cat" escalation ladder (#66): motion/last-seen/predicted
        # zoom crops -> YOLO -> VLM detect + confirm -> VLM query. On-demand only —
        # a Test-tool frame (works regardless of config) or a live camera's latest
        # frame (needs the vlm_escalation toggle). Never touches the treat path.
        import cv2

        body = request.get_json(silent=True) or {}
        settings = body.get("settings") or {}
        use_vlm = body.get("use_vlm", True) is not False
        camera = (body.get("camera") or "").strip()
        loop = app.config["loop"]

        det = None
        endpoint = None
        hints, tracks = [], []
        if camera:                          # --- live-camera path ---------------
            cfg = config_mod.load()
            if not cfg.vlm_escalation:
                return jsonify({"error": "Live-camera escalation is disabled — enable "
                                "'VLM escalation' in settings first."}), 403
            if not loop.is_running():
                return jsonify({"error": "Detection isn't running."}), 409
            det = loop.get_detector(camera)
            if det is None:
                return jsonify({"error": f"Unknown camera {camera!r}."}), 409
            raw = det.latest_frame()
            if raw is None:
                return jsonify({"error": "No frame from that camera yet."}), 409
            # Published frames are always already adjusted (0.59.0), on both the
            # grab-thread and synchronous paths — so never adjust here again.
            frame = det._crop(raw)
            # The trail endpoint first — where the last movement ENDED is the
            # strongest "look here" hint (#67) — then motion blobs + last sighting.
            endpoint = det.trail_endpoint()
            cam_zones = next((c.get("zones") or [] for c in (cfg.cameras or [])
                              if isinstance(c, dict) and c.get("name") == camera), [])
            if endpoint:
                # The precise exit check (#68): an endpoint inside a doorway zone
                # means "may have left the view", whatever the frame-edge test said.
                if box_in_exit_zone(endpoint["box"], cam_zones, det.roi):
                    endpoint = {**endpoint, "interior": False}
                hints.append(tuple(endpoint["box"]))
                tracks.append((endpoint["ts"], tuple(endpoint["box"])))
            hints += list(det.motion_hint_boxes())
            if det.motion_hint_boxes() and det.motion_hint_ts() > 0:
                tracks.append((det.motion_hint_ts(), det.motion_hint_boxes()[0]))
            recent = [s for s in loop.cats.recent(limit=20)
                      if s.get("camera") == camera][:3]
            if recent:
                hints.append(tuple(recent[0]["box"]))
            tracks += [(s["ts"], tuple(s["box"])) for s in recent]
        else:                               # --- Test-tool frame path -----------
            with _TEST_SESSIONS_LOCK:
                frames = _TEST_SESSIONS.get(body.get("id"))
            if not frames:
                return jsonify({"error": "Upload a frame first (it may have expired)."}), 404
            try:
                idx = max(0, min(int(body.get("frame_index", 0)), len(frames) - 1))
            except (TypeError, ValueError):
                idx = 0
            frame = frames[idx]
            # Optional client-supplied hint boxes (no motion data for an upload).
            for b in body.get("hints") or []:
                if isinstance(b, (list, tuple)) and len(b) == 4:
                    hints.append(tuple(int(v) for v in b))

        h, w = frame.shape[:2]
        pred = escalation.predict_hint_box(tracks, frame_size=(w, h))
        if pred:
            hints.append(pred)

        cat_conf = settings.get("cat_confidence")
        try:
            cat_conf = float(cat_conf) if cat_conf is not None else 0.5
        except (TypeError, ValueError):
            cat_conf = 0.5
        lc = settings.get("locator_classes")
        locator_classes = tuple(lc) if isinstance(lc, (list, tuple)) and lc else ("cat",)
        run_yolo = _yolo_boxes_fn(settings, floor=min(cat_conf, 0.3))

        vlm_detect = vlm_query = None
        vlm_note = None
        if use_vlm:
            mode = body.get("mode") or vlm.DEFAULT_MODE
            model = body.get("model") or vlm.DEFAULT_MODEL
            key = _vlm_key(body)
            passes = body.get("passes") or vlm.DEFAULT_PASSES
            try:
                vlm.preflight(key, mode)
            except RuntimeError as exc:
                return jsonify({"error": str(exc)}), 503
            vlm_detect = lambda img: vlm.detect_regions(   # noqa: E731
                img, "cat", model=model, api_key=key, mode=mode)["objects"]
            vlm_query = lambda img: vlm.query_image_voted(  # noqa: E731
                img, model=model, api_key=key, mode=mode, passes=passes)
        else:
            vlm_note = "VLM rungs disabled by request — zoom+yolo only."

        result = escalation.escalate(
            frame, hints, run_yolo=run_yolo, vlm_detect=vlm_detect,
            vlm_query=vlm_query, locator_classes=locator_classes,
            cat_confidence=cat_conf)

        # "Probable location": nothing YOLO-CONFIRMED, but there's an honest lead.
        # Two kinds, NEVER recorded as a sighting: a votes-only VLM "yes" (#69 —
        # 37–42% decoy FP; voting fixes variance, not bias, so it can't confirm),
        # or a trail whose last movement ended in-view (interior endpoint, #67).
        # An endpoint at the frame edge / in an exit zone means "may have exited"
        # → no trail claim.
        result["probable"] = None
        if not result["found"]:
            if result.get("vlm_probable"):
                vp = result["vlm_probable"]
                result["probable"] = {
                    "box": list(vp["box"]), "kind": "vlm", "ratio": vp.get("ratio"),
                    "note": (f"The VLM voted yes ({vp.get('ratio') or '?'}) but YOLO "
                             "could not confirm — treated as a lead only, not a "
                             "sighting (VLMs false-fire on cat-shaped decoys).")}
            elif endpoint and endpoint["interior"]:
                result["probable"] = {
                    "box": list(endpoint["box"]), "kind": "trail",
                    "age_s": endpoint["age_s"],
                    "note": ("No confirmed detection, but the last movement ended here "
                             f"{int(endpoint['age_s'])}s ago and no exit was seen — "
                             "probable location.")}

        # Annotate: thin gray crop windows, green final box (or orange probable).
        annotated = frame.copy()
        for c in result["crops"]:
            x1, y1, x2, y2 = c["box"]
            cv2.rectangle(annotated, (x1, y1), (x2, y2), (160, 160, 160), 1)
        if result["found"] and result["box"]:
            x1, y1, x2, y2 = result["box"]
            cv2.rectangle(annotated, (x1, y1), (x2, y2), (80, 220, 80), 2)
            cv2.putText(annotated, f"{result['label']} ({result['source']})",
                        (x1, max(14, y1 - 6)), cv2.FONT_HERSHEY_SIMPLEX, 0.5,
                        (80, 220, 80), 1, cv2.LINE_AA)
        elif result["probable"]:
            x1, y1, x2, y2 = result["probable"]["box"]
            cv2.rectangle(annotated, (x1, y1), (x2, y2), (0, 165, 255), 2)
            cv2.putText(annotated, f"probable ({result['probable']['kind']})",
                        (x1, max(14, y1 - 6)),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 165, 255), 1, cv2.LINE_AA)
        for c in result["crops"]:
            x1, y1, x2, y2 = c["box"]
            c["image"] = _orig_thumb_data_url(frame[y1:y2, x1:x2], width=320)

        # A live-camera find is a real sighting: record it (tagged with its rung)
        # and flash the feed via the existing boost — same flow as the still scan.
        if camera and det is not None and result["found"]:
            snap_jpeg = cv2.imencode(".jpg", annotated,
                                     [cv2.IMWRITE_JPEG_QUALITY, 80])[1].tobytes()
            snap = loop.snapshots.save(snap_jpeg)
            sighting = loop.cats.record(camera, result["box"], (w, h),
                                        result["score"], image=snap,
                                        label=result["label"] or "cat",
                                        source=result["source"],
                                        zone=zone_for(result["box"], cam_zones, det.roi))
            where = f" ({sighting['region']})" if sighting["region"] else ""
            loop.activity.add(
                "motion",
                f"🔍 Cat found by escalation ({result['source']}){where} on {camera}.",
                image=snap)
            loop.boost_detection(camera, 5.0)
        elif camera and result["probable"] and result["probable"]["kind"] == "vlm":
            # An unconfirmed VLM lead on a live camera: don't record it — hand it
            # to the detector that CAN confirm (boost forces YOLO onto the next
            # frames; a real cat there becomes a normal, recorded sighting).
            loop.activity.add(
                "motion",
                f"🔍 Escalation: VLM suspects a cat on {camera} (unconfirmed) — "
                "boosting detection so YOLO can confirm.")
            # Targeted (0.42.0): the lead names a PLACE — aim the boost's zoom
            # crop + heaviest-model pass at it, don't just look harder everywhere.
            loop.boost_detection(camera, 10.0, box=result["probable"]["box"])

        # Attach the trail image (the probable state's evidence; nice context on any
        # live run). Data URL so the GUI shows it inline next to the rung table.
        trail_url = None
        if det is not None:
            trail_img = det._trail.render(frame)
            if trail_img is not None:
                trail_url = _full_jpeg_data_url(trail_img, quality=80)

        return jsonify({**result, "annotated": _full_jpeg_data_url(annotated, quality=80),
                        "trail": trail_url,
                        "frame_size": [w, h], "hints_used": [list(b) for b in hints],
                        "camera": camera or None, "note": vlm_note})

    @app.post("/api/vlm/temporal")
    def api_vlm_temporal():
        # Temporal VLM check (#68): tile the last few frames (a camera's ring buffer,
        # or an uploaded video's sampled frames) into one numbered mosaic and ask the
        # voted question "did a cat appear or pass through?". moondream is image-only;
        # the mosaic is the honest 3070-friendly stand-in for video input — whether it
        # reasons well over grids is exactly what this tester measures.
        body = request.get_json(silent=True) or {}
        camera = (body.get("camera") or "").strip()
        loop = app.config["loop"]

        if camera:                          # live path: same privacy gate as escalate
            cfg = config_mod.load()
            if not cfg.vlm_escalation:
                return jsonify({"error": "Live-camera checks are disabled — enable "
                                "'VLM escalation' in settings first."}), 403
            if not loop.is_running():
                return jsonify({"error": "Detection isn't running."}), 409
            det = loop.get_detector(camera)
            if det is None:
                return jsonify({"error": f"Unknown camera {camera!r}."}), 409
            frames = det.recent_frames()
            if len(frames) < 2:
                return jsonify({"error": "Not enough frame history yet — give the "
                                "camera a few seconds."}), 409
        else:                               # uploaded video (or image) session
            with _TEST_SESSIONS_LOCK:
                sess = _TEST_SESSIONS.get(body.get("id"))
            if not sess:
                return jsonify({"error": "Upload a video first (it may have expired)."}), 404
            # Sampled frames carry no timestamps; index-spaced stand-ins label the grid.
            frames = [(float(i), f) for i, f in enumerate(sess)]

        mosaic = escalation.frame_mosaic(frames)
        if mosaic is None:
            return jsonify({"error": "No frames to build a mosaic from."}), 409
        try:
            vlm.preflight(_vlm_key(body), body.get("mode") or vlm.DEFAULT_MODE)
            result = vlm.query_image_voted(
                mosaic, prompt=vlm.TEMPORAL_PROMPT,
                model=body.get("model") or vlm.DEFAULT_MODEL,
                mode=body.get("mode") or vlm.DEFAULT_MODE,
                api_key=_vlm_key(body), passes=body.get("passes") or vlm.DEFAULT_PASSES)
        except RuntimeError as exc:
            return jsonify({"error": str(exc)}), 503
        span = round(frames[-1][0] - frames[0][0], 1)
        # A temporal "yes" is a HINT, not a verdict (#69: VLM decoy FP is high and
        # nothing here was YOLO-confirmed). On a live camera, act on the hint the
        # honest way: boost detection so YOLO looks hard at the next frames — a
        # real cat becomes a normal recorded sighting; a decoy dies quietly.
        note = None
        if result.get("answer") == "yes":
            note = "A temporal 'yes' is an unconfirmed hint, not a sighting."
            if camera:
                loop.activity.add(
                    "motion",
                    f"⏱️ Temporal check suspects a cat passed on {camera} "
                    "(unconfirmed) — boosting detection so YOLO can confirm.")
                loop.boost_detection(camera, 10.0)
                note += " Detection boosted on the camera so YOLO can confirm."
        return jsonify({**result, "mosaic": _full_jpeg_data_url(mosaic, quality=80),
                        "n_frames": min(len(frames), escalation.MOSAIC_MAX_TILES),
                        "span_s": span, "camera": camera or None, "hint_note": note})

    @app.get("/api/test/benchmark/<rid>.html")
    def api_benchmark_html(rid):
        rep = _find_report(rid)
        if not rep:
            return jsonify({"error": "report expired"}), 404
        slug = rep.get("slug") or f"benchmark_{rid}"
        return Response(rep["html"], mimetype="text/html",
                        headers={"Content-Disposition": f'inline; filename="{slug}.html"'})

    @app.get("/api/test/benchmark/<rid>.xlsx")
    def api_benchmark_xlsx(rid):
        rep = _find_report(rid)
        if not rep or not rep.get("xlsx"):
            return jsonify({"error": "no XLSX (install openpyxl) or report expired"}), 404
        slug = rep.get("slug") or f"benchmark_{rid}"
        return Response(
            rep["xlsx"],
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{slug}.xlsx"'})

    @app.get("/api/test/benchmark/<sid>/all.zip")
    def api_benchmark_zip(sid):
        # "Download all" (#35): one zip of the summary + every per-image report, each
        # named with its slug — unzip into a folder, host together, links resolve.
        import io
        import zipfile

        rep = _find_report(sid)
        members = (rep or {}).get("members")
        if not members:
            return jsonify({"error": "no batch bundle (single report or expired)"}), 404
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for slug, html in members:
                zf.writestr(f"{slug}.html", html)
        return Response(
            buf.getvalue(), mimetype="application/zip",
            headers={"Content-Disposition":
                     f'attachment; filename="{rep.get("slug") or "benchmark-batch"}.zip"'})

    @app.post("/api/cats/boost")
    def api_cats_boost():
        # "Show cat" jumped the live feed to this camera — run its detector
        # continuously for a short window so the feed boxes the cat while you look.
        name = (request.get_json(silent=True) or {}).get("camera")
        ok = app.config["loop"].boost_detection(name) if name else False
        return jsonify({"ok": bool(ok)})

    return app
